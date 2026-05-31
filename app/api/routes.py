import os
import re
import secrets
import shutil
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Request, Form, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import inspect, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.database import SessionLocal, engine
from app.models.candidate_model import Candidate
from app.models.candidate_event_model import CandidateEvent
from app.models.cv_analysis_model import CvAnalysis
from app.models.job_post_model import JobPost
from app.models.interview_model import Interview
from app.models.scheduling_token_model import SchedulingToken
from app.models.interview_scorecard_model import InterviewScorecard
from app.models.scorecard_token_model import ScorecardToken
from app.services.cv_parser import (
    process_cvs_for_job,
    extract_text_from_file,
    analyze_cv_with_ai,
    parse_ai_response,
    normalize_data,
    save_candidate_to_db,
    match_job_profile,
    build_fallback_profile,
    clean_score,
    recommendation_to_status,
    as_text,
    normalize_text,
    _is_domain_incompatible,
)
from app.services.cloudinary_service import upload_cv_to_cloudinary, stream_cv_from_cloudinary

from app.services.gmail_service import (
    download_cv_attachments,
    send_interview_email,
    send_rejection_email,
    send_email_api
)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# No-cache environment for public pages — avoids Jinja2 LRU cache
# incompatibility with newer Python/Jinja2 versions
import jinja2 as _jinja2
_public_env = _jinja2.Environment(
    loader=_jinja2.FileSystemLoader("app/templates"),
    autoescape=True,
    cache_size=0,
)

UPLOAD_FOLDER = "app/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
TERMINAL_STATUSES = ["REFUZAT", "ANGAJAT", "EXCLUS"]
ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".docx", ".doc"}


def safe_upload_filename(filename):
    filename = filename or f"cv_{int(time.time())}.pdf"
    filename = filename.replace("\\", "_").replace("/", "_")
    filename = re.sub(r"[^A-Za-z0-9._() -]+", "_", filename)
    filename = re.sub(r"\s+", " ", filename).strip()
    return filename or f"cv_{int(time.time())}.pdf"


def unique_upload_path(filename):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    filename = safe_upload_filename(filename)
    base, ext = os.path.splitext(filename)
    candidate = filename
    index = 1

    while os.path.exists(os.path.join(UPLOAD_FOLDER, candidate)):
        candidate = f"{base}_{index}{ext}"
        index += 1

    return os.path.join(UPLOAD_FOLDER, candidate), candidate


def is_allowed_cv_file(filename):
    ext = os.path.splitext(filename or "")[1].lower()
    return ext in ALLOWED_UPLOAD_EXTENSIONS


def ensure_candidate_extra_columns():
    if not inspect(engine).has_table("candidates"):
        return

    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS batch_id VARCHAR"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS visible_in_dashboard INTEGER DEFAULT 1"))
        conn.execute(text("UPDATE candidates SET visible_in_dashboard = 1 WHERE visible_in_dashboard IS NULL"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS companies TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS job_id INTEGER"))
        # ── Campuri noi agent AI decizie ──
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS decision_reason TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS missing_requirements TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS must_verify_by_phone TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS recommended_next_action VARCHAR"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS priority VARCHAR"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS manager_summary TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS phone_call_script TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS interview_questions TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS better_role_match VARCHAR"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS reject_reason_internal TEXT"))
        conn.execute(text("ALTER TABLE candidates ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"))


ensure_candidate_extra_columns()


def ensure_interview_extra_columns():
    if not inspect(engine).has_table("interviews"):
        return
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE interviews ADD COLUMN IF NOT EXISTS reminder_sent BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE interviews ADD COLUMN IF NOT EXISTS reminder_sent_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE interviews ADD COLUMN IF NOT EXISTS action_token VARCHAR"))


ensure_interview_extra_columns()


def ensure_scorecard_columns():
    if not inspect(engine).has_table("interview_scorecards"):
        return
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE interview_scorecards ADD COLUMN IF NOT EXISTS evaluator_name VARCHAR"))


ensure_scorecard_columns()


def ensure_scorecard_token_columns():
    if not inspect(engine).has_table("scorecard_tokens"):
        return
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE scorecard_tokens ADD COLUMN IF NOT EXISTS evaluator_email VARCHAR"))
        conn.execute(text("ALTER TABLE scorecard_tokens ADD COLUMN IF NOT EXISTS evaluator_name VARCHAR"))


ensure_scorecard_token_columns()


# ── Scorecard helpers ────────────────────────────────────────────────────────

VALID_RECOMMENDATIONS = ["Strong Hire", "Hire", "No Hire"]


def calculate_average_score(scorecard) -> float:
    scores = [
        scorecard.technical_skills_score,
        scorecard.attention_to_detail_score,
        scorecard.communication_score,
        scorecard.team_fit_score,
    ]
    valid = [s for s in scores if s is not None and 1 <= s <= 5]
    if not valid:
        return 0.0
    return round(sum(valid) / len(valid), 1)


def get_candidate_scorecard(candidate_id: int):
    db = SessionLocal()
    sc = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == candidate_id
    ).order_by(InterviewScorecard.id.desc()).first()
    db.close()
    return sc


def get_scorecard_by_interview(candidate_id: int, interview_id: int):
    db = SessionLocal()
    sc = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == candidate_id,
        InterviewScorecard.interview_id == interview_id,
    ).first()
    db.close()
    return sc


def get_candidate_interviews(candidate_id: int):
    db = SessionLocal()
    interviews = db.query(Interview).filter(
        Interview.candidate_id == candidate_id
    ).order_by(Interview.id.desc()).all()
    db.close()
    return interviews


def load_candidates():
    db = SessionLocal()
    candidates = db.query(Candidate).order_by(Candidate.id.desc()).all()
    total = len(candidates)
    db.close()
    return candidates, total


def load_dashboard_candidates():
    db = SessionLocal()
    candidates = db.query(Candidate).filter(
        Candidate.visible_in_dashboard == 1
    ).filter(
        Candidate.status.notin_(TERMINAL_STATUSES)
    ).order_by(
        Candidate.id.desc()
    ).all()
    total = len(candidates)
    db.close()
    return candidates, total


def get_candidate_by_id(candidate_id):
    db = SessionLocal()
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    db.close()
    return candidate


def get_candidate_events(candidate_id):
    db = SessionLocal()
    events = db.query(CandidateEvent).filter(
        CandidateEvent.candidate_id == candidate_id
    ).order_by(CandidateEvent.id.desc()).all()
    db.close()
    return events


def get_last_candidate_event(candidate_id):
    db = SessionLocal()
    event = db.query(CandidateEvent).filter(
        CandidateEvent.candidate_id == candidate_id
    ).order_by(CandidateEvent.id.desc()).first()
    db.close()
    return event


def log_candidate_event(candidate_id, event_type, title, description=""):
    db = SessionLocal()
    event = CandidateEvent(
        candidate_id=candidate_id,
        event_type=event_type,
        title=title,
        description=description
    )
    db.add(event)
    db.commit()
    db.close()


def get_max_candidate_id():
    db = SessionLocal()
    candidate = db.query(Candidate).order_by(Candidate.id.desc()).first()
    max_id = candidate.id if candidate else 0
    db.close()
    return max_id


def get_final_cv_files():
    db = SessionLocal()
    rows = db.query(Candidate).filter(
        Candidate.status.in_(TERMINAL_STATUSES)
    ).all()
    files = {row.cv_file for row in rows if row.cv_file}
    db.close()
    return files


def mark_new_candidates_batch(after_id, batch_id, visible=1, only_files=None, hide_final_files=True):
    final_files = get_final_cv_files() if hide_final_files else set()
    db = SessionLocal()
    rows = db.query(Candidate).filter(Candidate.id > after_id).all()

    for row in rows:
        row.batch_id = batch_id

        should_show = bool(visible)

        if only_files is not None and row.cv_file not in only_files:
            should_show = False

        if row.cv_file in final_files:
            should_show = False

        if row.status in TERMINAL_STATUSES:
            should_show = False

        row.visible_in_dashboard = 1 if should_show else 0

    db.commit()
    db.close()


def clear_dashboard_only():
    db = SessionLocal()
    rows = db.query(Candidate).all()

    for row in rows:
        row.visible_in_dashboard = 0

    db.commit()
    db.close()


def update_candidate_status(candidate_id, status, event_title=None, event_description=""):
    db = SessionLocal()
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()

    if not candidate:
        db.close()
        return None

    old_status = candidate.status or "NOU"
    candidate.status = status

    if status in TERMINAL_STATUSES:
        candidate.visible_in_dashboard = 0

    db.commit()
    db.refresh(candidate)
    db.close()

    log_candidate_event(
        candidate_id=candidate_id,
        event_type=status,
        title=event_title or f"Status schimbat: {status_display(old_status)} -> {status_display(status)}",
        description=event_description
    )

    return candidate


def update_candidate_manual(candidate_id, name, email, phone, status, position, experience, skills, summary):
    db = SessionLocal()
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()

    if not candidate:
        db.close()
        return None

    old_status = candidate.status or "NOU"

    candidate.name = name.strip() if name else candidate.name
    candidate.email = email.strip() if email else ""
    candidate.phone = phone.strip() if phone else ""
    candidate.status = status.strip() if status else candidate.status
    candidate.position = position.strip() if position else ""
    candidate.experience = experience.strip() if experience else ""
    candidate.skills = skills.strip() if skills else ""
    candidate.summary = summary.strip() if summary else ""

    if candidate.status in TERMINAL_STATUSES:
        candidate.visible_in_dashboard = 0

    new_status = candidate.status or "NOU"

    db.commit()
    db.refresh(candidate)
    db.close()

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="EDIT",
        title="Date candidat actualizate manual",
        description=f"Status anterior: {status_display(old_status)}. Status actual: {status_display(new_status)}."
    )

    return candidate


def group_candidates_by_job(candidates):
    jobs = {}

    for candidate in candidates:
        job_name = candidate.job_title or "Fara categorie"

        if job_name not in jobs:
            jobs[job_name] = {"latest_id": candidate.id, "candidates": []}

        jobs[job_name]["latest_id"] = max(jobs[job_name]["latest_id"], candidate.id)
        jobs[job_name]["candidates"].append(candidate)

    sorted_jobs = sorted(jobs.items(), key=lambda item: item[1]["latest_id"], reverse=True)

    final_jobs = {}
    for job_name, data in sorted_jobs:
        def _sort_key(c, jn=job_name):
            incompatibil = _is_domain_incompatible(c.position or "", jn)
            return (0 if incompatibil else 1, c.score or 0)
        final_jobs[job_name] = sorted(data["candidates"], key=_sort_key, reverse=True)

    return final_jobs


def extract_risk(summary, risk_name):
    text_value = (summary or "").lower()
    marker = risk_name.lower() + ":"

    if marker not in text_value:
        return "unknown"

    after = text_value.split(marker, 1)[1].strip()

    if after.startswith("low"):
        return "low"
    if after.startswith("medium"):
        return "medium"
    if after.startswith("high"):
        return "high"

    return "unknown"


def clean_summary(summary):
    value = summary or ""
    value = re.sub(
        r"\bRisc\s+supracalificare\s*:\s*(low|medium|high|unknown|scazut|mediu|ridicat|necunoscut)\s*\.?\s*",
        "",
        value,
        flags=re.IGNORECASE
    )
    value = re.sub(
        r"\bNepotrivire\s+nivel\s*:\s*(low|medium|high|unknown|scazut|mediu|ridicat|necunoscut)\s*\.?\s*",
        "",
        value,
        flags=re.IGNORECASE
    )
    value = re.sub(r"\s+", " ", value).strip()
    return value or "Fara rezumat"


def risk_label(value):
    value = (value or "unknown").lower()
    if value == "low":
        return "scazut"
    if value == "medium":
        return "mediu"
    if value == "high":
        return "ridicat"
    return "necunoscut"


def status_display(status):
    status = status or "NOU"

    mapping = {
        "ADMIS": "Top Candidat",
        "DE ANALIZAT": "Potential Candidat",
        "RESPINS": "Potrivit altor roluri",
        "SELECTAT": "Potential Candidat",
        "INTERVIU": "Interviu",
        "REFUZAT": "Potrivit altor roluri",
        "ANGAJAT": "Angajat",
        "EXCLUS": "Candidat Exclus",
        "NOU": "Nou"
    }

    return mapping.get(status, status)


def decision_text(candidate):
    score = candidate.score or 0
    status = candidate.status or "NOU"

    messages = {
        "INTERVIU": "Email trimis. Candidatul trebuie contactat pentru prima discutie.",
        "REFUZAT": "Candidatul a fost inchis pentru acest rol si poate ramane potrivit pentru alte roluri.",
        "EXCLUS": "Candidat exclus din proces. Nu se trimite email si nu mai apare in dashboard.",
        "ANGAJAT": "Candidatul a fost marcat ca angajat.",
        "SELECTAT": "Candidatul are potential si a fost selectat pentru urmatorul pas.",
        "ADMIS": "Top candidat pentru acest rol. Verifica detaliile si disponibilitatea.",
        "DE ANALIZAT": "Potential candidat. Necesita validare telefonica.",
        "RESPINS": "Candidatul pare mai potrivit pentru alte roluri."
    }

    if status in messages:
        return messages[status]

    if score >= 70:
        return "Profil cu potrivire buna."
    if score >= 45:
        return "Profil intermediar. Necesita verificare."
    return "Profil slab potrivit pentru acest rol."


def status_class(status):
    status = status or "NOU"

    if status in ["ADMIS", "ANGAJAT"]:
        return "green"
    if status in ["DE ANALIZAT", "SELECTAT", "INTERVIU"]:
        return "yellow"
    if status in ["RESPINS", "REFUZAT", "EXCLUS"]:
        return "red"
    return "default"


def event_class(event_type):
    event_type = event_type or "INFO"

    if event_type in ["EMAIL_INTERVIU", "ANGAJAT", "INTERVIU"]:
        return "green"
    if event_type in ["EMAIL_REFUZ", "REFUZAT", "EXCLUS"]:
        return "red"
    if event_type in ["STATUS", "EDIT", "SELECTAT", "NOTE", "DE ANALIZAT"]:
        return "yellow"
    return "default"


def format_event_date(value):
    if not value:
        return ""
    try:
        return value.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(value)


def calc_percent(part, total):
    if not total:
        return 0
    return round((part / total) * 100, 1)


def temporarily_keep_only_files(files_to_keep):
    hold_folder = os.path.join(UPLOAD_FOLDER, "_hidden_temp")
    os.makedirs(hold_folder, exist_ok=True)
    moved = []

    for filename in os.listdir(UPLOAD_FOLDER):
        full_path = os.path.join(UPLOAD_FOLDER, filename)

        if not os.path.isfile(full_path):
            continue

        if filename not in files_to_keep:
            target = os.path.join(hold_folder, filename)
            shutil.move(full_path, target)
            moved.append((target, full_path))

    return moved


def restore_temp_files(moved_files):
    for old_path, new_path in moved_files:
        if os.path.exists(old_path):
            shutil.move(old_path, new_path)


def build_analytics(candidates):
    total = len(candidates)
    status_counter = Counter(candidate.status or "NOU" for candidate in candidates)
    scores = [candidate.score or 0 for candidate in candidates]
    avg_score = round(sum(scores) / total, 1) if total else 0

    high_score = len([score for score in scores if score >= 70])
    medium_score = len([score for score in scores if 45 <= score < 70])
    low_score = len([score for score in scores if score < 45])

    job_stats = []
    grouped = defaultdict(list)

    for candidate in candidates:
        grouped[candidate.job_title or "Fara categorie"].append(candidate)

    for job_title, items in grouped.items():
        job_scores = [item.score or 0 for item in items]
        job_status = Counter(item.status or "NOU" for item in items)

        job_stats.append({
            "job_title": job_title,
            "total": len(items),
            "avg_score": round(sum(job_scores) / len(items), 1) if items else 0,
            "admis": job_status.get("ADMIS", 0),
            "de_analizat": job_status.get("DE ANALIZAT", 0),
            "interviu": job_status.get("INTERVIU", 0),
            "selectat": job_status.get("SELECTAT", 0),
            "refuzat": job_status.get("REFUZAT", 0),
            "exclus": job_status.get("EXCLUS", 0),
            "angajat": job_status.get("ANGAJAT", 0),
            "respins": job_status.get("RESPINS", 0),
            "top_candidate": max(items, key=lambda item: item.score or 0) if items else None
        })

    job_stats = sorted(job_stats, key=lambda item: item["total"], reverse=True)
    top_candidates = sorted(candidates, key=lambda candidate: candidate.score or 0, reverse=True)[:10]

    pipeline = [
        {"label": "Nou", "count": status_counter.get("NOU", 0)},
        {"label": "Potential Candidat", "count": status_counter.get("DE ANALIZAT", 0) + status_counter.get("SELECTAT", 0)},
        {"label": "Interviu", "count": status_counter.get("INTERVIU", 0)},
        {"label": "Top Candidat", "count": status_counter.get("ADMIS", 0)},
        {"label": "Angajat", "count": status_counter.get("ANGAJAT", 0)},
        {"label": "Potrivit altor roluri", "count": status_counter.get("REFUZAT", 0) + status_counter.get("RESPINS", 0)},
        {"label": "Candidat Exclus", "count": status_counter.get("EXCLUS", 0)}
    ]

    return {
        "total": total,
        "avg_score": avg_score,
        "high_score": high_score,
        "medium_score": medium_score,
        "low_score": low_score,
        "status_counter": status_counter,
        "job_stats": job_stats,
        "top_candidates": top_candidates,
        "pipeline": pipeline,
        "rates": {
            "interview_rate": calc_percent(status_counter.get("INTERVIU", 0), total),
            "hire_rate": calc_percent(status_counter.get("ANGAJAT", 0), total),
            "reject_rate": calc_percent(
                status_counter.get("REFUZAT", 0) + status_counter.get("RESPINS", 0) + status_counter.get("EXCLUS", 0),
                total
            ),
            "shortlist_rate": calc_percent(
                status_counter.get("ADMIS", 0)
                + status_counter.get("DE ANALIZAT", 0)
                + status_counter.get("SELECTAT", 0)
                + status_counter.get("INTERVIU", 0)
                + status_counter.get("ANGAJAT", 0),
                total
            )
        }
    }


def apply_candidate_filters(candidates, q="", status="", job="", min_score=0):
    filtered = candidates

    if q:
        search = q.lower().strip()
        filtered = [
            candidate for candidate in filtered
            if search in (candidate.name or "").lower()
            or search in (candidate.email or "").lower()
            or search in (candidate.phone or "").lower()
            or search in (candidate.cv_file or "").lower()
        ]

    if status:
        filtered = [candidate for candidate in filtered if (candidate.status or "") == status]

    if job:
        filtered = [candidate for candidate in filtered if (candidate.job_title or "") == job]

    if min_score:
        filtered = [candidate for candidate in filtered if (candidate.score or 0) >= min_score]

    return filtered


def candidate_identity_key(candidate):
    """Grupeaza acelasi om dupa email, telefon sau fisier CV."""
    email_value = (candidate.email or "").strip().lower()
    phone_value = re.sub(r"\D+", "", candidate.phone or "")
    cv_value = (candidate.cv_file or "").strip().lower()

    if email_value:
        return f"email:{email_value}"
    if phone_value:
        return f"phone:{phone_value}"
    if cv_value:
        return f"cv:{cv_value}"
    return f"candidate:{candidate.id}"


def candidate_search_blob(candidate):
    parts = [
        candidate.name,
        candidate.email,
        candidate.phone,
        candidate.cv_file,
        candidate.job_title,
        candidate.position,
        candidate.companies,
        candidate.experience,
        candidate.skills,
        candidate.summary,
        candidate.strengths,
        candidate.weaknesses,
        candidate.level,
    ]
    return " ".join([str(part) for part in parts if part]).lower()


def build_unique_candidate_rows(candidates, q="", status="", job="", skill="", min_score=0):
    """
    Intoarce o lista unica de candidati.
    Fiecare rand contine candidatul principal si toate potrivirile pe posturi.
    """
    grouped = defaultdict(list)
    for candidate in candidates:
        grouped[candidate_identity_key(candidate)].append(candidate)

    rows = []
    q_value = (q or "").strip().lower()
    status_value = (status or "").strip()
    job_value = (job or "").strip().lower()
    skill_value = (skill or "").strip().lower()

    for _, items in grouped.items():
        sorted_items = sorted(items, key=lambda item: (item.score or 0, item.id or 0), reverse=True)
        best = sorted_items[0]

        if status_value and not any((item.status or "") == status_value for item in sorted_items):
            continue

        # Filtru job: candidatul trebuie sa aiba cel putin o analiza pentru jobul cautat
        if job_value and not any(
            job_value in (item.job_title or "").lower()
            or job_value in (item.position or "").lower()
            for item in sorted_items
        ):
            continue

        if skill_value and not any(skill_value in candidate_search_blob(item) for item in sorted_items):
            continue

        if q_value and not any(q_value in candidate_search_blob(item) for item in sorted_items):
            continue

        # Cand filtrul de job e activ, scorul afisat = scorul pe jobul cautat (nu best overall)
        if job_value:
            job_items = [item for item in sorted_items
                         if job_value in (item.job_title or "").lower()
                         or job_value in (item.position or "").lower()]
            job_items_sorted = sorted(job_items, key=lambda x: x.score or 0, reverse=True)
            display_best = job_items_sorted[0] if job_items_sorted else best
            display_score = display_best.score or 0
        else:
            display_best = best
            display_score = best.score or 0

        if min_score and display_score < min_score:
            continue

        # Construieste lista de matches, cu jobul cautat primul daca filtrul e activ
        matches = []
        seen_jobs = set()
        all_items_ordered = sorted_items

        if job_value:
            # Punem jobul cautat primul, restul dupa scor
            job_first = [i for i in sorted_items if job_value in (i.job_title or "").lower() or job_value in (i.position or "").lower()]
            rest = [i for i in sorted_items if i not in job_first]
            all_items_ordered = job_first + rest

        for item in all_items_ordered:
            job_name = item.job_title or item.position or "Nespecificat"
            key = (job_name or "").strip().lower()
            if key in seen_jobs:
                continue
            seen_jobs.add(key)
            matches.append({
                "id": item.id,
                "job_title": job_name,
                "position": item.position or "Nespecificat",
                "score": item.score or 0,
                "status": item.status or "NOU",
                "summary": clean_summary(item.summary),
                "cv_file": item.cv_file,
            })

        rows.append({
            "candidate": best,
            "best_match": display_best,
            "matches": matches,
            "best_score": display_score,
            "match_count": len(matches),
            "emails": sorted({item.email for item in sorted_items if item.email}),
            "phones": sorted({item.phone for item in sorted_items if item.phone}),
        })

    # Daca filtrul de job e activ, candidatii incompatibili de domeniu merg la coada
    if job_value:
        def _row_sort(row, jv=job_value):
            nat = row["candidate"].position or ""
            incompatibil = _is_domain_incompatible(nat, jv)
            return (0 if incompatibil else 1, row["best_score"])
        rows = sorted(rows, key=_row_sort, reverse=True)
    else:
        rows = sorted(rows, key=lambda row: (row["best_score"], row["candidate"].id or 0), reverse=True)
    return rows


def candidate_last_event_text(candidate_id):
    event = get_last_candidate_event(candidate_id)

    if not event:
        return "", ""

    return event.title or "", format_event_date(event.created_at)


def generate_candidates_excel(candidates: list) -> BytesIO:
    """Export candidati in Excel cu toate campurile AI, color-coded dupa scor."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidati"

    headers = [
        "Nume", "Email", "Telefon", "Job cautat", "Pozitie detectata",
        "Scor", "Status", "Prioritate", "Experienta", "Skill-uri",
        "Motiv Decizie", "Cerinte Lipsa", "Verificare Telefon",
        "Script Apel", "Intrebari Interviu", "Rol Potrivit Altul",
        "Motiv Refuz Intern", "Rezumat AI", "CV fisier",
        "Ultima actiune", "Data ultimei actiuni",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_green  = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_orange = PatternFill("solid", fgColor="F8CBAD")

    for candidate in candidates:
        last_action, last_action_date = candidate_last_event_text(candidate.id)
        score = candidate.score or 0
        row_fill = fill_green if score >= 85 else (fill_yellow if score >= 70 else fill_orange)

        ws.append([
            candidate.name or "",
            candidate.email or "",
            candidate.phone or "",
            candidate.job_title or "",
            candidate.position or "",
            score,
            status_display(candidate.status),
            candidate.priority or "",
            candidate.experience or "",
            candidate.skills or "",
            candidate.decision_reason or "",
            candidate.missing_requirements or "",
            candidate.must_verify_by_phone or "",
            candidate.phone_call_script or "",
            candidate.interview_questions or "",
            candidate.better_role_match or "",
            candidate.reject_reason_internal or "",
            candidate.summary or "",
            candidate.cv_file or "",
            last_action,
            last_action_date,
        ])

        for cell in ws[ws.max_row]:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 25, "B": 32, "C": 16, "D": 25, "E": 25,
        "F": 8,  "G": 20, "H": 12, "I": 16, "J": 40,
        "K": 35, "L": 30, "M": 25, "N": 40, "O": 40,
        "P": 25, "Q": 30, "R": 60, "S": 30, "T": 28, "U": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


templates.env.globals["extract_risk"] = extract_risk
templates.env.globals["risk_label"] = risk_label
templates.env.globals["clean_summary"] = clean_summary
templates.env.globals["decision_text"] = decision_text
templates.env.globals["status_class"] = status_class
templates.env.globals["status_display"] = status_display
templates.env.globals["event_class"] = event_class
templates.env.globals["format_event_date"] = format_event_date
templates.env.globals["calc_percent"] = calc_percent


@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, background_tasks: BackgroundTasks):
    candidates, total = load_dashboard_candidates()
    jobs = group_candidates_by_job(candidates)
    message = request.query_params.get("message")

    # ── Remindere interviuri (background, nu blocheaza pagina) ────────────────
    base_url = str(request.base_url).rstrip("/")
    _schedule_interview_reminders(background_tasks, base_url=base_url)

    # ── Agentul recomanda azi ──────────────────────────────────────────────────
    call_now   = sorted([c for c in candidates if (c.recommended_next_action or "") == "CALL_NOW"],
                        key=lambda x: x.score or 0, reverse=True)
    call_later = sorted([c for c in candidates if (c.recommended_next_action or "") == "CALL_LATER"],
                        key=lambda x: x.score or 0, reverse=True)
    keep_other = [c for c in candidates if (c.recommended_next_action or "") == "KEEP_FOR_OTHER_ROLE"]
    auto_reject= [c for c in candidates if (c.recommended_next_action or "") == "REJECT"]
    top3       = (call_now + call_later)[:3]

    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context={
            "request": request,
            "candidates": candidates,
            "jobs": jobs,
            "total": total,
            "message": message,
            "call_now": call_now,
            "call_later": call_later,
            "keep_other": keep_other,
            "auto_reject": auto_reject,
            "top3": top3,
        }
    )


@router.get("/candidati", response_class=HTMLResponse)
async def candidates_page(request: Request, q: str = "", status: str = "", job: str = "", skill: str = "", min_score: int = 0, page: int = 1):
    candidates, total = load_candidates()
    all_rows = build_unique_candidate_rows(candidates, q=q, status=status, job=job, skill=skill, min_score=min_score)

    all_jobs = sorted(list({candidate.job_title for candidate in candidates if candidate.job_title}))
    all_statuses = sorted(list({candidate.status for candidate in candidates if candidate.status}))

    # Paginare
    PAGE_SIZE = 50
    total_pages = max(1, (len(all_rows) + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(1, min(page, total_pages))
    candidate_rows = all_rows[(page - 1) * PAGE_SIZE : page * PAGE_SIZE]

    # Toate randurile unice (fara filtru) — pentru unique_total
    all_unique = build_unique_candidate_rows(candidates)

    stats = {
        "total": len(candidates),
        "unique_total": len(all_unique),
        "filtered": len(all_rows),
        "analyses_filtered": sum(len(row["matches"]) for row in all_rows),
        "interviu": len([c for c in candidates if c.status == "INTERVIU"]),
        "selectat": len([c for c in candidates if c.status == "SELECTAT"]),
        "angajat": len([c for c in candidates if c.status == "ANGAJAT"]),
        "refuzat": len([c for c in candidates if c.status == "REFUZAT"]),
        "exclus": len([c for c in candidates if c.status == "EXCLUS"]),
        "de_analizat": len([c for c in candidates if c.status == "DE ANALIZAT"]),
    }

    # Scorecard map: candidate_id -> scorecard (cel mai recent)
    db = SessionLocal()
    all_scorecards = db.query(InterviewScorecard).all()
    db.close()
    sc_by_candidate = {}
    for sc in all_scorecards:
        if sc.candidate_id not in sc_by_candidate:
            sc_by_candidate[sc.candidate_id] = sc

    return templates.TemplateResponse(
        request=request,
        name="candidates.html",
        context={
            "request": request,
            "candidates": [row["candidate"] for row in candidate_rows],
            "candidate_rows": candidate_rows,
            "total": total,
            "stats": stats,
            "all_jobs": all_jobs,
            "all_statuses": all_statuses,
            "q": q,
            "status": status,
            "job": job,
            "skill": skill,
            "min_score": min_score,
            "page": page,
            "total_pages": total_pages,
            "sc_by_candidate": sc_by_candidate,
        }
    )


@router.get("/candidati/export")
async def export_candidates(q: str = "", status: str = "", job: str = "", skill: str = "", min_score: int = 0):
    candidates, total = load_candidates()
    rows = build_unique_candidate_rows(candidates, q=q, status=status, job=job, skill=skill, min_score=min_score)
    filtered = [row["candidate"] for row in rows]

    output = generate_candidates_excel(filtered)
    filename = f"nexas_hr_candidati_{int(time.time())}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request):
    candidates, total = load_candidates()
    analytics = build_analytics(candidates)

    return templates.TemplateResponse(
        request=request,
        name="analytics.html",
        context={"request": request, "candidates": candidates, "total": total, "analytics": analytics}
    )


@router.get("/candidat/{candidate_id}", response_class=HTMLResponse)
async def candidate_detail(request: Request, candidate_id: int):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate:
        return templates.TemplateResponse(
            request=request,
            name="candidate_detail.html",
            context={"request": request, "candidate": None, "events": [], "message": "Candidatul nu a fost gasit."},
            status_code=404
        )

    message = request.query_params.get("message")
    error = request.query_params.get("error")
    events = get_candidate_events(candidate_id)
    interviews = get_candidate_interviews(candidate_id)

    db = SessionLocal()
    scorecards = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == candidate_id
    ).all()
    db.close()

    scorecard_map = {sc.interview_id: sc for sc in scorecards}
    latest_scorecard = scorecards[-1] if scorecards else None
    avg_score = calculate_average_score(latest_scorecard) if latest_scorecard else None

    return templates.TemplateResponse(
        request=request,
        name="candidate_detail.html",
        context={
            "request": request,
            "candidate": candidate,
            "events": events,
            "message": message,
            "error": error,
            "interviews": interviews,
            "scorecard_map": scorecard_map,
            "avg_score": avg_score,
        }
    )


@router.post("/candidat/{candidate_id}/salveaza")
async def save_candidate_manual(
    candidate_id: int,
    name: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    status: str = Form(""),
    position: str = Form(""),
    experience: str = Form(""),
    skills: str = Form(""),
    summary: str = Form("")
):
    candidate = update_candidate_manual(candidate_id, name, email, phone, status, position, experience, skills, summary)

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    return RedirectResponse(
        url=f"/candidat/{candidate_id}?message=Datele candidatului au fost actualizate.",
        status_code=303
    )


@router.post("/candidat/{candidate_id}/nota")
async def add_candidate_note(candidate_id: int, note: str = Form("")):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    note = note.strip()

    if not note:
        return RedirectResponse(
            url=f"/candidat/{candidate_id}?error=Nota HR este goala.",
            status_code=303
        )

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="NOTE",
        title="Nota HR adaugata",
        description=note
    )

    return RedirectResponse(
        url=f"/candidat/{candidate_id}?message=Nota HR a fost salvata.",
        status_code=303
    )


@router.get("/cv/{candidate_id}")
async def open_candidate_cv(candidate_id: int):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate or not candidate.cv_file:
        return JSONResponse({"error": "CV-ul nu a fost gasit pentru acest candidat."}, status_code=404)

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="CV_DESCĂRCAT",
        title="CV descărcat",
        description=f"CV accesat: {candidate.cv_file}",
    )

    # Incearca din Cloudinary (persistent, nu depinde de Render disk)
    buffer, mime = stream_cv_from_cloudinary(candidate.cv_file)
    if buffer:
        return StreamingResponse(
            buffer,
            media_type=mime,
            headers={
                "Content-Disposition": f'inline; filename="{candidate.cv_file}"'
            }
        )

    # Fallback: disk local (doar in dev)
    file_path = os.path.join("app", "uploads", candidate.cv_file)
    if os.path.exists(file_path):
        return FileResponse(path=file_path, filename=candidate.cv_file, media_type="application/pdf")

    return JSONResponse({"error": "CV-ul nu a fost gasit."}, status_code=404)


@router.get("/candidat/{candidate_id}/trimite-interviu")
async def redirect_to_calendar_for_interview(candidate_id: int):
    """Redirecteaza catre pagina calendar cu candidatul preselectat."""
    return RedirectResponse(url=f"/calendar?candidate_id={candidate_id}", status_code=303)


@router.post("/candidat/{candidate_id}/trimite-interviu")
async def redirect_to_calendar_for_interview_post(candidate_id: int):
    """Redirecteaza catre pagina calendar cu candidatul preselectat (POST fallback)."""
    return RedirectResponse(url=f"/calendar?candidate_id={candidate_id}", status_code=303)


@router.post("/candidat/{candidate_id}/refuza")
async def reject_candidate(candidate_id: int):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    if not candidate.email:
        update_candidate_status(
            candidate_id,
            "REFUZAT",
            event_title="Candidat mutat la Potrivit altor roluri fara email",
            event_description="Candidatul nu are email extras din CV. Nu s-a trimis email."
        )
        return RedirectResponse(
            url=f"/candidat/{candidate_id}?message=Status actualizat: Potrivit altor roluri. Candidatul nu are email extras din CV, deci nu s-a trimis email.",
            status_code=303
        )

    try:
        send_rejection_email(to_email=candidate.email, candidate_name=candidate.name, job_title=candidate.job_title)
        update_candidate_status(
            candidate_id,
            "REFUZAT",
            event_title="Email de refuz trimis",
            event_description=f"Email trimis catre {candidate.email}."
        )

        return RedirectResponse(
            url=f"/candidat/{candidate_id}?message=Email de refuz trimis. Status actualizat: Potrivit altor roluri.",
            status_code=303
        )

    except Exception as e:
        return RedirectResponse(
            url=f"/candidat/{candidate_id}?error=Eroare la trimiterea emailului de refuz: {str(e)}",
            status_code=303
        )


@router.post("/candidat/{candidate_id}/exclude")
async def exclude_candidate(candidate_id: int, reason: str = Form("")):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    reason = reason.strip() or "Candidat exclus manual din proces. Nu s-a trimis email."

    update_candidate_status(
        candidate_id,
        "EXCLUS",
        event_title="Candidat exclus",
        event_description=reason
    )

    return RedirectResponse(
        url=f"/candidat/{candidate_id}?message=Candidatul a fost exclus. Nu s-a trimis email.",
        status_code=303
    )


@router.post("/candidat/{candidate_id}/angajeaza")
async def hire_candidate(candidate_id: int):
    candidate = update_candidate_status(
        candidate_id,
        "ANGAJAT",
        event_title="Candidat marcat ca angajat",
        event_description="Decizie finala pozitiva."
    )

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    return RedirectResponse(url=f"/candidat/{candidate_id}?message=Status actualizat: Angajat.", status_code=303)


@router.post("/candidat/{candidate_id}/selecteaza")
async def select_candidate(candidate_id: int):
    candidate = update_candidate_status(
        candidate_id,
        "SELECTAT",
        event_title="Potential candidat selectat",
        event_description="Candidatul a fost selectat manual pentru urmatorul pas."
    )

    if not candidate:
        return RedirectResponse(url="/?message=Candidatul nu a fost gasit.", status_code=303)

    return RedirectResponse(url=f"/candidat/{candidate_id}?message=Status actualizat: Potential Candidat.", status_code=303)


@router.post("/analizeaza-job")
async def analyze_job(target_job: str = Form(...)):
    batch_id = str(int(time.time()))
    after_id = get_max_candidate_id()
    final_files = get_final_cv_files()

    # Cauta daca exista un post in platforma cu cerinte definite
    job_requirements = None
    try:
        db = SessionLocal()
        job_posts = db.query(JobPost).filter(JobPost.activ == True).all()
        db.close()
        job_norm = normalize_text(target_job)
        best_match, best_score = None, 0
        for post in job_posts:
            post_norm = normalize_text(post.titlu or "")
            if post_norm == job_norm:
                score = 100
            elif post_norm in job_norm or job_norm in post_norm:
                score = 70
            else:
                pw = set(post_norm.split())
                jw = set(job_norm.split())
                score = int(len(pw & jw) / max(len(pw), 1) * 45) if pw else 0
            if score > best_score:
                best_match, best_score = post, score
        if best_match and best_score >= 30:
            reqs = " ".join(filter(None, [
                best_match.cerinte or "",
                best_match.responsabilitati or "",
            ])).strip()
            if reqs:
                job_requirements = reqs
                print(f"[ANALIZA] Post gasit in platforma: '{best_match.titlu}' — cerinte folosite pentru scoring.")
    except Exception as e:
        print(f"[ANALIZA] Nu s-au putut incarca cerintele din platforma: {e}")

    result = process_cvs_for_job(target_job, job_requirements=job_requirements)

    mark_new_candidates_batch(
        after_id=after_id,
        batch_id=batch_id,
        visible=1,
        only_files=None,
        hide_final_files=True
    )

    message = result.get("message", "Analiza finalizata.")
    if final_files:
        message += " Candidatii marcati Angajat, Exclus sau Potrivit altor roluri nu mai apar in dashboard."

    return RedirectResponse(url=f"/?message={message}", status_code=303)


@router.post("/upload-cv-uri")
async def upload_cv_files(cv_files: list[UploadFile] = File(...)):
    """
    Incarca manual CV-uri in app/uploads, fara analiza imediata.
    Analiza se face ulterior din butonul "Analizeaza existente".
    Asa poti incarca CV-uri azi, de mai multe ori, si cauta joburi maine.
    """
    uploaded_files = []
    skipped_files = []

    for upload in cv_files:
        original_name = safe_upload_filename(upload.filename)

        if not is_allowed_cv_file(original_name):
            skipped_files.append(original_name)
            continue

        content = await upload.read()

        if not content:
            skipped_files.append(original_name)
            continue

        file_path, final_name = unique_upload_path(original_name)

        with open(file_path, "wb") as output_file:
            output_file.write(content)

        upload_cv_to_cloudinary(file_path, final_name)

        uploaded_files.append(final_name)

    if not uploaded_files:
        return RedirectResponse(
            url="/?message=Nu s-a incarcat niciun CV valid. Acceptam PDF, DOCX si DOC.",
            status_code=303
        )

    message = (
        f"CV-uri incarcate manual: {len(uploaded_files)}. "
        f"Fisiere ignorate: {len(skipped_files)}. "
        "CV-urile au fost salvate. Le poti analiza ulterior din butonul Analizeaza existente."
    )

    return RedirectResponse(url=f"/?message={message}", status_code=303)


@router.post("/descarca-gmail")
async def descarca_gmail():
    """
    Descarca CV-uri noi din Gmail fara a face analiza.
    Duplicatele sunt verificate pe Cloudinary - nu re-descarca CV-uri existente.
    """
    gmail_result = download_cv_attachments(max_results=30)

    message = (
        f"Gmail verificat. "
        f"CV-uri noi descarcate: {gmail_result['downloaded_count']}. "
        f"CV-uri deja existente (sarite): {gmail_result['skipped_count']}. "
        f"Acum poti rula analiza pentru jobul dorit."
    )

    return RedirectResponse(
        url=f"/?message={message}",
        status_code=303
    )


@router.post("/gmail-si-analiza")
async def gmail_and_analyze(target_job: str = Form("")):
    """
    Descarca CV-uri noi din Gmail + analizeaza pentru jobul specificat.
    Duplicatele sunt verificate pe Cloudinary.
    """
    gmail_result = download_cv_attachments(max_results=30)

    downloaded = gmail_result["downloaded_count"]
    skipped = gmail_result["skipped_count"]

    if target_job.strip() and downloaded > 0:
        try:
            analysis_result = process_cvs_for_job(target_job.strip())
            message = (
                f"Gmail verificat. CV-uri noi: {downloaded}, sarite: {skipped}. "
                f"Analiza finalizata: {analysis_result.get('saved', 0)} candidati salvati pentru {target_job}."
            )
        except Exception as e:
            message = (
                f"Gmail verificat. CV-uri noi: {downloaded}, sarite: {skipped}. "
                f"Eroare la analiza: {str(e)}"
            )
    elif target_job.strip() and downloaded == 0:
        try:
            analysis_result = process_cvs_for_job(target_job.strip())
            message = (
                f"Gmail verificat. Nu exista CV-uri noi (toate deja descarcate). "
                f"Analiza din cache: {analysis_result.get('saved', 0)} candidati pentru {target_job}."
            )
        except Exception as e:
            message = f"Gmail verificat. CV-uri noi: 0. Eroare la analiza: {str(e)}"
    else:
        message = (
            f"Gmail verificat. CV-uri noi descarcate: {downloaded}, sarite: {skipped}. "
            f"Nu ai specificat un job pentru analiza."
        )

    return RedirectResponse(url=f"/?message={message}", status_code=303)


@router.post("/sterge-baza")
async def clear_database():
    clear_dashboard_only()
    return RedirectResponse(
        url="/?message=Dashboard curatat. Candidatii raman salvati in pagina Candidati.",
        status_code=303
    )


@router.post("/sterge-definitiv")
async def delete_all_database(confirm_delete: str = Form("")):
    if confirm_delete != "STERGE":
        return RedirectResponse(url="/?message=Stergere anulata. Token de confirmare invalid.", status_code=303)
    db = SessionLocal()
    db.query(CandidateEvent).delete()
    db.query(Candidate).delete()
    db.commit()
    db.close()

    return RedirectResponse(url="/?message=Toata baza de date a fost stearsa definitiv.", status_code=303)


# ─── RE-ANALIZĂ CU LOGICĂ NOUĂ ──────────────────────────────────────────────

_reanalize_progress = {"total": 0, "done": 0, "running": False}


def _run_reanalizare_completa():
    global _reanalize_progress
    try:
        # 1. Sterge cache-ul
        db = SessionLocal()
        db.query(CvAnalysis).delete()
        db.commit()

        # 2. Colecteaza perechi unice (cv_file, job_title)
        candidates = db.query(Candidate).filter(
            Candidate.cv_file != None,
            Candidate.job_title != None,
        ).all()
        db.close()

        pairs = {}
        for c in candidates:
            key = (c.cv_file.strip(), c.job_title.strip())
            if key[0] and key[1]:
                pairs[key] = True

        unique_pairs = list(pairs.keys())
        _reanalize_progress["total"] = len(unique_pairs)
        _reanalize_progress["done"] = 0
        _reanalize_progress["running"] = True

        print(f"[RE-ANALIZARE] Start: {len(unique_pairs)} perechi unice CV + job")

        for cv_file, job_title in unique_pairs:
            try:
                path = os.path.join(UPLOAD_FOLDER, cv_file)
                if not os.path.exists(path):
                    _reanalize_progress["done"] += 1
                    continue

                from app.services.cv_parser import extract_text_from_file, analyze_cv_with_ai, parse_ai_response, normalize_data, save_analysis_to_cache, match_job_profile, build_fallback_profile
                text = extract_text_from_file(path)
                if not text:
                    _reanalize_progress["done"] += 1
                    continue

                ai_response = analyze_cv_with_ai(text, job_title)
                data = parse_ai_response(ai_response)
                if data:
                    profile = match_job_profile(job_title) or build_fallback_profile(job_title)
                    normalized = normalize_data(data, cv_file, job_title, text, profile)
                    save_analysis_to_cache(cv_file, job_title, normalized)
                    print(f"[RE-ANALIZARE] ✓ {cv_file} → {job_title}")
            except Exception as e:
                print(f"[RE-ANALIZARE] ✗ {cv_file}: {e}")
            finally:
                _reanalize_progress["done"] += 1

    except Exception as e:
        print(f"[RE-ANALIZARE] Eroare generala: {e}")
    finally:
        _reanalize_progress["running"] = False
        print(f"[RE-ANALIZARE] Finalizat: {_reanalize_progress['done']}/{_reanalize_progress['total']}")


@router.post("/reanalize-completa")
async def reanalize_completa(background_tasks: BackgroundTasks):
    global _reanalize_progress
    if _reanalize_progress.get("running"):
        return JSONResponse({"success": False, "message": "Re-analizarea este deja in curs."})

    background_tasks.add_task(_run_reanalizare_completa)

    db = SessionLocal()
    total_cv = db.query(Candidate).filter(
        Candidate.cv_file != None,
        Candidate.job_title != None,
    ).count()
    db.close()

    return JSONResponse({
        "success": True,
        "message": f"Re-analizare pornita pentru ~{total_cv} candidati. Procesul ruleaza in fundal.",
        "total": total_cv,
    })


@router.get("/reanalize-status")
async def reanalize_status():
    p = _reanalize_progress
    return JSONResponse({
        "running": p.get("running", False),
        "total": p.get("total", 0),
        "done": p.get("done", 0),
        "percent": int(p["done"] / p["total"] * 100) if p.get("total") else 0,
    })


# ─── POSTURI VACANTE (INTERN) ────────────────────────────────────────────────

@router.get("/posturi", response_class=HTMLResponse)
async def jobs_page(request: Request, message: str = ""):
    db = SessionLocal()
    jobs = db.query(JobPost).order_by(JobPost.created_at.desc()).all()
    # Numar aplicatii per post
    applicant_counts = {}
    for job in jobs:
        count = db.query(Candidate).filter(Candidate.job_id == job.id).count()
        applicant_counts[job.id] = count
    db.close()
    return templates.TemplateResponse(
        request=request,
        name="jobs.html",
        context={"request": request, "jobs": jobs, "message": message, "applicant_counts": applicant_counts}
    )


@router.post("/posturi/adauga")
async def add_job(
    titlu: str = Form(...),
    departament: str = Form(""),
    locatie: str = Form(""),
    tip_contract: str = Form("Full-time"),
    descriere: str = Form(""),
    responsabilitati: str = Form(""),
    cerinte: str = Form(""),
    ce_oferim: str = Form("")
):
    db = SessionLocal()
    job = JobPost(
        titlu=titlu,
        departament=departament,
        locatie=locatie,
        tip_contract=tip_contract,
        descriere=descriere,
        responsabilitati=responsabilitati,
        cerinte=cerinte,
        ce_oferim=ce_oferim,
        activ=True
    )
    db.add(job)
    db.commit()
    db.close()
    return RedirectResponse(url="/posturi?message=Post adaugat cu succes.", status_code=303)


@router.post("/posturi/{job_id}/sterge")
async def delete_job(job_id: int):
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id).first()
    if job:
        db.delete(job)
        db.commit()
    db.close()
    return RedirectResponse(url="/posturi?message=Post sters.", status_code=303)


@router.post("/posturi/{job_id}/toggle")
async def toggle_job(job_id: int):
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id).first()
    status_msg = "inactiv"
    if job:
        job.activ = not job.activ
        db.commit()
        # Citim valoarea inainte sa inchidem sesiunea
        status_msg = "activ" if job.activ else "inactiv"
    db.close()
    return RedirectResponse(url=f"/posturi?message=Post marcat ca {status_msg}.", status_code=303)


@router.get("/posturi/{job_id}/aplicatii", response_class=HTMLResponse)
async def job_applicants(request: Request, job_id: int):
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id).first()
    if not job:
        db.close()
        return HTMLResponse("<h2>Post negasit.</h2>", status_code=404)

    # Exclude candidatii fara analiza AI (scor 0 si summary default)
    candidates = db.query(Candidate).filter(
        Candidate.job_id == job_id,
        Candidate.score > 0
    ).order_by(Candidate.score.desc()).all()
    db.close()

    top = [c for c in candidates if (c.score or 0) >= 75]
    potential = [c for c in candidates if 50 <= (c.score or 0) < 75]
    rest = [c for c in candidates if (c.score or 0) < 50]

    return templates.TemplateResponse(
        request=request,
        name="job_applicants.html",
        context={
            "request": request,
            "job": job,
            "top": top,
            "potential": potential,
            "rest": rest,
            "total": len(candidates),
            "status_display": status_display
        }
    )


@router.post("/posturi/{job_id}/editeaza")
async def edit_job(
    job_id: int,
    titlu: str = Form(...),
    departament: str = Form(""),
    locatie: str = Form(""),
    tip_contract: str = Form("Full-time"),
    descriere: str = Form(""),
    responsabilitati: str = Form(""),
    cerinte: str = Form(""),
    ce_oferim: str = Form("")
):
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id).first()
    if job:
        job.titlu = titlu
        job.departament = departament
        job.locatie = locatie
        job.tip_contract = tip_contract
        job.descriere = descriere
        job.responsabilitati = responsabilitati
        job.cerinte = cerinte
        job.ce_oferim = ce_oferim
        db.commit()
    db.close()
    return RedirectResponse(url="/posturi?message=Post actualizat.", status_code=303)


# ─── PAGINA PUBLICA CARIERE ───────────────────────────────────────────────────

@router.get("/cariere", response_class=HTMLResponse)
async def careers_page(request: Request):
    db = SessionLocal()
    jobs = db.query(JobPost).filter(JobPost.activ == True).order_by(JobPost.created_at.desc()).all()
    db.close()
    return templates.TemplateResponse(
        request=request,
        name="careers.html",
        context={"request": request, "jobs": jobs}
    )


@router.get("/cariere/{job_id}", response_class=HTMLResponse)
async def career_detail_page(request: Request, job_id: int):
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id, JobPost.activ == True).first()
    db.close()
    if not job:
        return HTMLResponse("<h2>Post indisponibil.</h2>", status_code=404)
    return templates.TemplateResponse(
        request=request,
        name="career_detail.html",
        context={"request": request, "job": job}
    )


def _background_analyze_and_save(
    cv_bytes: bytes,
    final_filename: str,
    local_path: str,
    job_title: str,
    job_id: int,
    job_locatie: str,
    nume: str,
    email: str,
    telefon: str
):
    """
    Ruleaza in background dupa ce Ion a primit confirmarea.
    1. Analizeaza CV-ul cu AI
    2. Salveaza candidatul in DB cu scor real
    3. Trimite emailuri
    Zero blocare pe server.
    """
    try:
        # Scrie fisierul local temporar
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        with open(local_path, "wb") as f:
            f.write(cv_bytes)

        # Extrage text din CV
        cv_text = extract_text_from_file(local_path)

        if cv_text:
            # Analiza AI
            profile = match_job_profile(job_title) or build_fallback_profile(job_title)
            ai_response = analyze_cv_with_ai(cv_text, job_title)
            data = parse_ai_response(ai_response)

            if data:
                normalized = normalize_data(data, final_filename, job_title, cv_text, profile)
                score = clean_score(normalized.get("score", 0))
                status = recommendation_to_status(normalized.get("recommendation", "CONSIDER"))

                # Salveaza candidat cu scor real in DB
                db = SessionLocal()
                candidate = Candidate(
                    name=as_text(normalized.get("name", nume)),
                    email=as_text(normalized.get("email", email)) or email,
                    phone=as_text(normalized.get("phone", telefon)) or telefon,
                    position=as_text(normalized.get("position", job_title)),
                    experience=as_text(normalized.get("years_experience", "")),
                    skills=as_text(normalized.get("skills", "")),
                    companies=as_text(normalized.get("companies", "")),
                    score=score,
                    level=as_text(normalized.get("level", "")),
                    strengths=as_text(normalized.get("strengths", "")),
                    weaknesses=as_text(normalized.get("weaknesses", "")),
                    summary=as_text(normalized.get("summary", "")),
                    job_title=job_title,
                    job_id=job_id,
                    status=status,
                    cv_file=final_filename
                )
                db.add(candidate)
                db.commit()
                candidate_id = candidate.id
                db.close()

                log_candidate_event(
                    candidate_id=candidate_id,
                    event_type="aplicatie",
                    title=f"Aplicatie directa — {job_title}",
                    description=f"CV analizat AI. Scor: {score}. Status: {status}."
                )

                # Notificare HR cu scor real
                hr_email = os.getenv("GMAIL_EMAIL", "")
                if hr_email:
                    try:
                        send_email_api(
                            to_email=hr_email,
                            subject=f"Aplicatie noua: {normalized.get('name', nume)} — {job_title} (scor: {score})",
                            body=f"""Aplicatie noua primita pe NEXAS HR.

Candidat: {normalized.get('name', nume)}
Email: {email}
Telefon: {telefon or 'necompletat'}
Post: {job_title}
Locatie: {job_locatie or 'nespecificata'}
Scor AI: {score}/100
Status: {status}

Vezi candidatul: https://hr-g66k.onrender.com/candidat/{candidate_id}
"""
                        )
                    except Exception:
                        pass
        else:
            # CV fara text (imagine scanned) — salvam fara scor AI
            db = SessionLocal()
            candidate = Candidate(
                name=nume, email=email, phone=telefon,
                job_title=job_title, job_id=job_id,
                position=job_title, status="NOU", score=0,
                cv_file=final_filename,
                summary=f"CV primit. Nu s-a putut extrage text (posibil scanat)."
            )
            db.add(candidate)
            db.commit()
            db.close()

    except Exception as e:
        print(f"EROARE background analiza {final_filename}: {e}")
    finally:
        # Sterge fisierul local dupa analiza
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
        except Exception:
            pass

    # Email confirmare catre candidat (se trimite dupa analiza)
    try:
        send_email_api(
            to_email=email,
            subject=f"CV-ul tau pentru postul {job_title} a fost primit",
            body=f"""Buna ziua, {nume},

Va multumim pentru interesul acordat postului de {job_title}{' in ' + job_locatie if job_locatie else ''}.

CV-ul dumneavoastra a fost primit si analizat.

Vom reveni cu un raspuns in cel mai scurt timp posibil daca profilul dumneavoastra corespunde cerintelor postului.

Cu respect,
Echipa HR NEXAS
"""
        )
    except Exception:
        pass


def _make_candidate_visible_as_nou(candidate_id, summary_note):
    """Helper: daca analiza esueaza, candidatul devine vizibil cu status NOU."""
    try:
        db = SessionLocal()
        c = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if c:
            c.status = "NOU"
            c.visible_in_dashboard = 1
            c.summary = summary_note
            db.commit()
        db.close()
    except Exception as e:
        print(f"[BG] WARN make_visible error candidat {candidate_id}: {e}")


def _process_cv_background(
    candidate_id: int,
    local_path: str,
    final_filename: str,
    cv_bytes: bytes,
    job_id: int,
    job_titlu: str,
    job_locatie: str,
    nume: str,
    email: str,
    telefon: str,
):
    """
    Ruleaza in background dupa ce utilizatorul a primit deja confirmare.
    Face: upload Cloudinary + extragere text + analiza AI + update DB + emailuri.
    Orice eroare face candidatul vizibil ca NOU — recruiterul poate verifica manual.
    """
    from app.services.cloudinary_service import upload_cv_bytes_to_cloudinary

    # ── Email confirmare catre candidat — trimis IMEDIAT, indiferent de analiza ──
    try:
        send_email_api(
            to_email=email,
            subject=f"CV-ul tau pentru postul {job_titlu} a fost primit",
            body=f"""Buna ziua, {nume},

Va multumim pentru interesul acordat postului de {job_titlu}{' in ' + job_locatie if job_locatie else ''}.

CV-ul dumneavoastra a fost primit si va fi analizat in cel mai scurt timp.

Vom reveni cu un raspuns daca profilul dumneavoastra corespunde cerintelor postului.

Cu respect,
Echipa HR NEXAS
"""
        )
    except Exception as e:
        print(f"[BG] WARN email confirmare candidat esuat: {e}")

    # 1. Salveaza CV pe disc
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        with open(local_path, "wb") as f:
            f.write(cv_bytes)
    except Exception as e:
        print(f"[BG] EROARE scriere fisier {final_filename}: {e}")
        _make_candidate_visible_as_nou(candidate_id, "CV primit. Eroare la salvarea fisierului. Verifica manual.")
        return

    # 2. Upload Cloudinary
    try:
        upload_cv_bytes_to_cloudinary(cv_bytes, final_filename)
    except Exception as e:
        print(f"[BG] WARN Cloudinary upload esuat {final_filename}: {e}")

    # 3. Extrage text din CV
    try:
        cv_text = extract_text_from_file(local_path)
    except Exception as e:
        print(f"[BG] EROARE extragere text {final_filename}: {e}")
        cv_text = None

    if not cv_text or len(cv_text.strip()) < 80:
        print(f"[BG] CV fara text suficient {candidate_id} — facut vizibil ca NOU")
        _make_candidate_visible_as_nou(
            candidate_id,
            "CV primit dar nu s-a putut extrage text (posibil scanat sau protejat). Verifica manual CV-ul."
        )
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
        except Exception:
            pass
        return

    # 4. Analiza AI
    try:
        profile = match_job_profile(job_titlu) or build_fallback_profile(job_titlu)
        ai_response = analyze_cv_with_ai(cv_text, job_titlu)
        data = parse_ai_response(ai_response)
    except Exception as e:
        print(f"[BG] EROARE analiza AI {final_filename}: {e}")
        data = None

    if not data:
        print(f"[BG] AI nu a returnat date valide pentru {final_filename} — facut vizibil ca NOU")
        _make_candidate_visible_as_nou(
            candidate_id,
            "CV primit. Analiza AI nu a putut fi completata. Verifica manual CV-ul."
        )
        return

    normalized = normalize_data(data, final_filename, job_titlu, cv_text, profile)
    score = clean_score(normalized.get("score", 0))
    status = recommendation_to_status(normalized.get("recommendation", "CONSIDER"))

    if score <= 0:
        print(f"[BG] Scor 0 pentru {final_filename} — facut vizibil ca NOU")
        _make_candidate_visible_as_nou(
            candidate_id,
            "CV primit. Scor AI = 0. Verifica manual daca CV-ul este valid pentru acest post."
        )
        return

    # 5. Actualizeaza candidatul in DB cu datele AI
    try:
        db = SessionLocal()
        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if candidate:
            candidate.name = as_text(normalized.get("name", nume)) or nume
            candidate.email = as_text(normalized.get("email", email)) or email
            candidate.phone = as_text(normalized.get("phone", telefon)) or telefon
            candidate.position = as_text(normalized.get("position", job_titlu))
            candidate.experience = as_text(normalized.get("years_experience", ""))
            candidate.skills = as_text(normalized.get("skills", ""))
            candidate.companies = as_text(normalized.get("companies", ""))
            candidate.score = score
            candidate.level = as_text(normalized.get("level", ""))
            candidate.strengths = as_text(normalized.get("strengths", ""))
            candidate.weaknesses = as_text(normalized.get("weaknesses", ""))
            candidate.summary = as_text(normalized.get("summary", ""))
            candidate.status = status
            candidate.visible_in_dashboard = 1
            candidate.decision_reason         = as_text(normalized.get("decision_reason", ""))
            candidate.missing_requirements    = as_text(normalized.get("missing_requirements", ""))
            candidate.must_verify_by_phone    = as_text(normalized.get("must_verify_by_phone", ""))
            candidate.recommended_next_action = as_text(normalized.get("recommended_next_action", ""))
            candidate.priority                = as_text(normalized.get("priority", ""))
            candidate.manager_summary         = as_text(normalized.get("manager_summary", ""))
            candidate.phone_call_script       = as_text(normalized.get("phone_call_script", ""))
            candidate.interview_questions     = as_text(normalized.get("interview_questions", ""))
            candidate.better_role_match       = as_text(normalized.get("better_role_match", ""))
            candidate.reject_reason_internal  = as_text(normalized.get("reject_reason_internal", ""))
            db.commit()
        db.close()
    except Exception as e:
        print(f"[BG] EROARE update DB candidat {candidate_id}: {e}")
        return

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="aplicatie",
        title=f"Aplicatie directa - {job_titlu}",
        description=f"CV analizat AI. Scor: {score}. Status: {status}."
    )

    # 6. Trimite emailuri
    hr_email = os.getenv("GMAIL_EMAIL", "")
    if hr_email:
        try:
            send_email_api(
                to_email=hr_email,
                subject=f"Aplicatie noua: {as_text(normalized.get('name', nume)) or nume} - {job_titlu} (scor: {score})",
                body=f"""Aplicatie noua primita pe NEXAS HR.

Candidat: {as_text(normalized.get('name', nume)) or nume}
Email: {as_text(normalized.get('email', email)) or email}
Telefon: {as_text(normalized.get('phone', telefon)) or telefon or 'necompletat'}
Post: {job_titlu}
Locatie: {job_locatie or 'nespecificata'}
Scor AI: {score}/100
Status: {status}

Vezi candidatul: https://hr-g66k.onrender.com/candidat/{candidate_id}
"""
            )
        except Exception as e:
            print(f"[BG] WARN email HR esuat: {e}")



# ─── REANALIZA FORTATA ────────────────────────────────────────────────────────

def _reanalyze_cv_background(candidate_id: int, cv_file: str, job_titlu: str):
    """
    Sterge cache-ul pentru acest CV+job si reanalizeaza de la zero.
    Ruleaza in background — recruiterul primeste redirect instant.
    """
    # 1. Sterge din cache
    key = normalize_text(job_titlu)
    try:
        db = SessionLocal()
        db.query(CvAnalysis).filter(
            CvAnalysis.cv_file == cv_file,
            CvAnalysis.job_title_normalized == key
        ).delete()
        db.commit()
        db.close()
    except Exception as e:
        print(f"[REANALIZ] WARN stergere cache: {e}")

    # 2. Obtine CV-ul (Cloudinary primul, local fallback)
    local_path = os.path.join(UPLOAD_FOLDER, cv_file)
    try:
        buffer, mime = stream_cv_from_cloudinary(cv_file)
        if buffer:
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            with open(local_path, "wb") as f:
                f.write(buffer.read())
    except Exception as e:
        print(f"[REANALIZ] WARN Cloudinary fetch: {e}")

    if not os.path.exists(local_path):
        print(f"[REANALIZ] CV negasit local si pe Cloudinary: {cv_file}")
        return

    # 3. Extrage text
    cv_text = extract_text_from_file(local_path)
    if not cv_text or len(cv_text.strip()) < 80:
        print(f"[REANALIZ] CV fara text suficient: {cv_file}")
        return

    # 4. Analizeaza
    try:
        profile = match_job_profile(job_titlu) or build_fallback_profile(job_titlu)
        ai_response = analyze_cv_with_ai(cv_text, job_titlu)
        data = parse_ai_response(ai_response)
    except Exception as e:
        print(f"[REANALIZ] EROARE AI: {e}")
        return

    if not data:
        return

    normalized = normalize_data(data, cv_file, job_titlu, cv_text, profile)
    score = clean_score(normalized.get("score", 0))
    status = recommendation_to_status(normalized.get("recommendation", "CONSIDER"))

    # 5. Actualizeaza candidatul in DB
    try:
        db = SessionLocal()
        candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
        if candidate:
            candidate.position               = as_text(normalized.get("position", ""))
            candidate.experience             = as_text(normalized.get("years_experience", ""))
            candidate.skills                 = as_text(normalized.get("skills", ""))
            candidate.companies              = as_text(normalized.get("companies", ""))
            candidate.score                  = score
            candidate.level                  = as_text(normalized.get("level", ""))
            candidate.strengths              = as_text(normalized.get("strengths", ""))
            candidate.weaknesses             = as_text(normalized.get("weaknesses", ""))
            candidate.summary                = as_text(normalized.get("summary", ""))
            candidate.status                 = status
            candidate.decision_reason        = as_text(normalized.get("decision_reason", ""))
            candidate.missing_requirements   = as_text(normalized.get("missing_requirements", ""))
            candidate.must_verify_by_phone   = as_text(normalized.get("must_verify_by_phone", ""))
            candidate.recommended_next_action= as_text(normalized.get("recommended_next_action", ""))
            candidate.priority               = as_text(normalized.get("priority", ""))
            candidate.manager_summary        = as_text(normalized.get("manager_summary", ""))
            candidate.phone_call_script      = as_text(normalized.get("phone_call_script", ""))
            candidate.interview_questions    = as_text(normalized.get("interview_questions", ""))
            candidate.better_role_match      = as_text(normalized.get("better_role_match", ""))
            candidate.reject_reason_internal = as_text(normalized.get("reject_reason_internal", ""))
            db.commit()
        db.close()
    except Exception as e:
        print(f"[REANALIZ] EROARE DB: {e}")
        return

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="REANALIZA",
        title="CV reanalizat fortat",
        description=f"Cache sters si reanalizat. Scor nou: {score}. Status: {status}."
    )
    print(f"[REANALIZ] Done: {cv_file} → scor {score}, status {status}")


@router.post("/candidat/{candidate_id}/reanalizeaza")
async def force_reanalyze(candidate_id: int, background_tasks: BackgroundTasks):
    candidate = get_candidate_by_id(candidate_id)

    if not candidate:
        return RedirectResponse(url="/candidati?message=Candidatul nu a fost gasit.", status_code=303)
    if not candidate.cv_file or not candidate.job_title:
        return RedirectResponse(
            url=f"/candidat/{candidate_id}?error=CV-ul sau postul lipsesc, nu se poate reanaliza.",
            status_code=303
        )

    background_tasks.add_task(
        _reanalyze_cv_background,
        candidate_id=candidate_id,
        cv_file=candidate.cv_file,
        job_titlu=candidate.job_title,
    )

    return RedirectResponse(
        url=f"/candidat/{candidate_id}?message=Reanaliza pornita in fundal. Reincarca pagina in 30 de secunde.",
        status_code=303
    )


# ─── REMINDER INTERVIURI ──────────────────────────────────────────────────────

def _send_interview_reminder_email(idata: dict):
    """Trimite email reminder cu o zi inainte de interviu (la ora 19:00 ora Romaniei)."""
    from datetime import datetime as dt
    try:
        base_url = idata.get("base_url", "")
        action_token = idata.get("action_token", "")
        locatie_line = f"\nLocatie: {idata['locatie']}" if idata.get("locatie") else ""

        confirm_link = f"{base_url}/interviu/{action_token}/confirma" if action_token else ""
        cancel_link = f"{base_url}/interviu/{action_token}/anuleaza" if action_token else ""

        confirm_section = ""
        if confirm_link:
            confirm_section = f"""
------------------------------------------------------------
Confirma participarea:
{confirm_link}

Anuleaza interviul:
{cancel_link}
------------------------------------------------------------
"""

        send_email_api(
            to_email=idata["candidate_email"],
            subject=f"Reminder: interviu maine — {idata['job_title'] or 'postul aplicat'}",
            body=f"""Buna ziua, {idata['candidate_name'] or 'Candidat'},

Va reamintim ca maine, {idata['data']}, la ora {idata['ora']}, aveti programat un interviu pentru postul de {idata['job_title'] or 'postul aplicat'}.{locatie_line}
{confirm_section}
Cu respect,
Echipa HR NEXAS
""",
        )
    except Exception as e:
        print(f"[REMINDER] WARN email esuat interviu {idata['id']}: {e}")

    try:
        db = SessionLocal()
        i = db.query(Interview).filter(Interview.id == idata["id"]).first()
        if i:
            i.reminder_sent = True
            i.reminder_sent_at = dt.utcnow()
            db.commit()
        db.close()
    except Exception as e:
        print(f"[REMINDER] WARN update DB esuat: {e}")


def _schedule_interview_reminders(background_tasks: BackgroundTasks, base_url: str = ""):
    """Verifica interviuri de maine si trimite reminder dupa ora 19:00 ora Romaniei."""
    from datetime import datetime, timedelta
    # Ora Romaniei: UTC+3 vara (EEST), UTC+2 iarna — folosim +3 ca aproximare sigura
    ro_now = datetime.utcnow() + timedelta(hours=3)
    if ro_now.hour < 19:
        return  # Trimitem reminderele doar dupa ora 19:00 ora Romaniei
    tomorrow_ro = (ro_now + timedelta(days=1)).strftime("%Y-%m-%d")

    try:
        db = SessionLocal()
        interviews_maine = db.query(Interview).filter(
            Interview.data == tomorrow_ro,
            Interview.reminder_sent.isnot(True),
            Interview.status.in_(["programat", "confirmat"]),
            Interview.candidate_email.isnot(None),
        ).all()

        interviews_data = []
        for i in interviews_maine:
            try:
                if not i.action_token:
                    i.action_token = secrets.token_urlsafe(24)
                    db.commit()
                interviews_data.append({
                    "id": i.id,
                    "candidate_email": i.candidate_email,
                    "candidate_name": i.candidate_name,
                    "job_title": i.job_title,
                    "data": i.data,
                    "ora": i.ora,
                    "locatie": i.locatie,
                    "action_token": i.action_token,
                    "base_url": base_url,
                })
            except Exception:
                continue
        db.close()

        for idata in interviews_data:
            background_tasks.add_task(_send_interview_reminder_email, idata)
        if interviews_data:
            print(f"[REMINDER] Trimise {len(interviews_data)} remindere pentru maine {tomorrow_ro}")
    except Exception as e:
        print(f"[REMINDER] WARN schedule error: {e}")


# ─── ACTIVITATE RECENTA ───────────────────────────────────────────────────────

@router.get("/activitate", response_class=HTMLResponse)
async def activitate_page(request: Request, tip: str = "", zile: int = 30):
    from datetime import datetime, timedelta
    cutoff = datetime.utcnow() - timedelta(days=zile)

    db = SessionLocal()
    query = db.query(CandidateEvent).filter(CandidateEvent.created_at >= cutoff)
    if tip:
        query = query.filter(CandidateEvent.event_type == tip)
    events = query.order_by(CandidateEvent.created_at.desc()).limit(300).all()

    candidate_ids = list({e.candidate_id for e in events})
    candidates_map = {}
    if candidate_ids:
        cands = db.query(Candidate).filter(Candidate.id.in_(candidate_ids)).all()
        candidates_map = {c.id: c for c in cands}
    db.close()

    event_rows = [{"event": e, "candidate": candidates_map.get(e.candidate_id)} for e in events]

    all_types = [
        "aplicatie", "REANALIZA",
        "INTERVIU", "EMAIL_INTERVIU", "SELECTAT", "ANGAJAT",
        "REFUZAT", "EXCLUS", "EMAIL_REFUZ",
        "NOTE", "EDIT", "STATUS",
    ]

    return templates.TemplateResponse(
        request=request,
        name="activitate.html",
        context={
            "request": request,
            "event_rows": event_rows,
            "tip": tip,
            "zile": zile,
            "all_types": all_types,
            "total": len(event_rows),
        }
    )


@router.post("/cariere/{job_id}/aplica")
async def apply_to_job(
    background_tasks: BackgroundTasks,
    job_id: int,
    nume: str = Form(...),
    email: str = Form(...),
    telefon: str = Form(""),
    cv_file: UploadFile = File(...)
):
    # ── Validari rapide (fara I/O blocant) ──────────────────────────────────
    db = SessionLocal()
    job = db.query(JobPost).filter(JobPost.id == job_id, JobPost.activ == True).first()
    job_titlu = job.titlu if job else None
    job_locatie = job.locatie if job else None
    db.close()

    if not job_titlu:
        return HTMLResponse("<h2>Post indisponibil.</h2>", status_code=404)

    # ── Validare email ───────────────────────────────────────────────────────
    if not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email or ""):
        return RedirectResponse(url=f"/cariere/{job_id}?error=Adresa de email invalida.", status_code=303)

    # ── Validare telefon (optional, dar daca e completat trebuie sa fie valid) ──
    if telefon and not re.match(r'^\+?[0-9]{10,13}$', telefon.replace(" ", "").replace("-", "")):
        return RedirectResponse(url=f"/cariere/{job_id}?error=Numar de telefon invalid (10-13 cifre).", status_code=303)

    # ── Validare extensie fisier ─────────────────────────────────────────────
    ext = os.path.splitext(cv_file.filename or "")[1].lower()
    if ext not in {".pdf", ".docx", ".doc"}:
        return RedirectResponse(url=f"/cariere/{job_id}?error=Doar PDF, DOC, DOCX.", status_code=303)

    # Citim bytes-urile acum (singurul await necesar — rapida, nu blochează)
    cv_bytes = await cv_file.read()
    if not cv_bytes:
        return RedirectResponse(url=f"/cariere/{job_id}?error=Fisierul CV este gol.", status_code=303)

    # ── Validare dimensiune fisier (max 10 MB) ───────────────────────────────
    if len(cv_bytes) > 10 * 1024 * 1024:
        return RedirectResponse(url=f"/cariere/{job_id}?error=CV-ul este prea mare (maxim 10 MB).", status_code=303)

    # ── Verificare aplicatie duplicata (aceeasi adresa email, acelasi job, in ultimele 30 de zile) ──
    from datetime import datetime, timedelta
    db_check = SessionLocal()
    cutoff_date = datetime.utcnow() - timedelta(days=30)
    duplicate = db_check.query(Candidate).filter(
        Candidate.email == email,
        Candidate.job_id == job_id,
        Candidate.created_at >= cutoff_date,
    ).first()
    db_check.close()
    if duplicate:
        return RedirectResponse(
            url=f"/cariere/{job_id}?error=Ai aplicat deja pentru acest post in ultimele 30 de zile.",
            status_code=303,
        )

    safe_name = re.sub(r"[^A-Za-z0-9._() -]+", "_", cv_file.filename or "cv")
    final_filename = f"aplicatie_{job_id}_{int(time.time())}_{safe_name}"
    local_path = os.path.join(UPLOAD_FOLDER, final_filename)

    # ── Salvam candidatul imediat in DB cu datele din formular ──────────────
    # score=0 si status="pending" — vor fi actualizate de background task
    db = SessionLocal()
    candidate = Candidate(
        name=nume,
        email=email,
        phone=telefon,
        position=job_titlu,
        job_title=job_titlu,
        job_id=job_id,
        score=0,
        status="pending",
        cv_file=final_filename,
        visible_in_dashboard=0,  # ascuns pana cand AI-ul termina analiza
    )
    db.add(candidate)
    db.commit()
    candidate_id = candidate.id
    db.close()

    # ── Programam toata munca grea in background ─────────────────────────────
    # Utilizatorul primeste redirect INSTANT, fara sa astepte AI-ul
    background_tasks.add_task(
        _process_cv_background,
        candidate_id=candidate_id,
        local_path=local_path,
        final_filename=final_filename,
        cv_bytes=cv_bytes,
        job_id=job_id,
        job_titlu=job_titlu,
        job_locatie=job_locatie,
        nume=nume,
        email=email,
        telefon=telefon,
    )

    return RedirectResponse(url=f"/cariere/{job_id}?success=1", status_code=303)



# ─── CALENDAR INTERVIURI ──────────────────────────────────────────────────────

@router.get("/calendar", response_class=HTMLResponse)
async def calendar_page(request: Request, candidate_id: int = None):
    preselected = None
    if candidate_id:
        preselected = get_candidate_by_id(candidate_id)
    return templates.TemplateResponse(
        request=request,
        name="calendar.html",
        context={"request": request, "preselected_candidate": preselected}
    )


@router.get("/calendar/api/week")
async def get_week_interviews(start: str):
    """
    Returneaza JSON cu toate interviurile dintr-o saptamana.
    start = "YYYY-MM-DD" (luni)
    """
    from datetime import datetime, timedelta
    try:
        start_dt = datetime.strptime(start, "%Y-%m-%d")
        end_dt = start_dt + timedelta(days=6)
        end_str = end_dt.strftime("%Y-%m-%d")
    except ValueError:
        return JSONResponse({"error": "Format data invalid. Foloseste YYYY-MM-DD."}, status_code=400)

    db = SessionLocal()
    interviews = db.query(Interview).filter(
        Interview.data >= start,
        Interview.data <= end_str
    ).order_by(Interview.data, Interview.ora).all()
    db.close()

    return JSONResponse([{
        "id": i.id,
        "candidate_id": i.candidate_id,
        "candidate_name": i.candidate_name or "—",
        "job_title": i.job_title or "—",
        "data": i.data,
        "ora": i.ora,
        "locatie": i.locatie or "",
        "durata": i.durata or 60,
        "status": i.status or "programat",
    } for i in interviews])


@router.get("/calendar/api/candidati-search")
async def search_candidates_for_calendar(q: str = ""):
    """Autocomplete candidati pentru modalul de programare."""
    db = SessionLocal()
    q_lower = f"%{q.lower()}%"
    candidates = db.query(Candidate).filter(
        Candidate.visible_in_dashboard == 1,
        (Candidate.name.ilike(q_lower)) | (Candidate.email.ilike(q_lower))
    ).limit(10).all()
    db.close()
    return JSONResponse([{
        "id": c.id,
        "name": c.name or "—",
        "email": c.email or "",
        "job_title": c.job_title or "",
    } for c in candidates])


@router.post("/calendar/api/interview")
async def create_interview_api(
    background_tasks: BackgroundTasks,
    candidate_id: int = Form(...),
    data: str = Form(...),
    ora: str = Form(...),
    locatie: str = Form(""),
    durata: int = Form(60),
):
    """Salveaza interviul in DB si trimite email candidatului."""
    candidate = get_candidate_by_id(candidate_id)
    if not candidate:
        return JSONResponse({"error": "Candidatul nu a fost gasit."}, status_code=404)

    db = SessionLocal()
    interview = Interview(
        candidate_id=candidate_id,
        candidate_name=candidate.name,
        candidate_email=candidate.email,
        job_title=candidate.job_title,
        data=data,
        ora=ora,
        locatie=locatie,
        durata=durata,
        status="programat",
        action_token=secrets.token_urlsafe(24),
    )
    db.add(interview)
    db.commit()
    interview_id = interview.id
    db.close()

    # Actualizeaza statusul candidatului
    update_candidate_status(
        candidate_id,
        "INTERVIU",
        event_title="Interviu programat",
        event_description=f"Interviu programat pe {data} la ora {ora}. Locatie: {locatie or 'nespecificata'}."
    )

    # Trimite email candidat in background
    if candidate.email:
        background_tasks.add_task(
            _send_interview_scheduled_email,
            to_email=candidate.email,
            candidate_name=candidate.name or "Candidat",
            job_title=candidate.job_title or "postul pentru care ati aplicat",
            data=data,
            ora=ora,
            locatie=locatie,
        )

    # Notifica HR
    background_tasks.add_task(
        _notify_hr_interview_event,
        "programat_hr",
        candidate.name or "Candidat",
        candidate.job_title or "postul aplicat",
        data,
        ora,
        locatie,
        candidate.email or "",
    )

    return JSONResponse({
        "success": True,
        "interview": {
            "id": interview_id,
            "candidate_id": candidate_id,
            "candidate_name": candidate.name or "—",
            "job_title": candidate.job_title or "—",
            "data": data,
            "ora": ora,
            "locatie": locatie,
            "durata": durata,
            "status": "programat",
        }
    })


@router.post("/calendar/api/interview/{interview_id}/sterge")
async def delete_interview_api(interview_id: int):
    """Sterge un interviu programat."""
    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        db.close()
        return JSONResponse({"error": "Interviul nu a fost gasit."}, status_code=404)
    db.delete(interview)
    db.commit()
    db.close()
    return JSONResponse({"success": True})


@router.post("/calendar/api/interview/{interview_id}/anuleaza")
async def cancel_interview_api(interview_id: int, background_tasks: BackgroundTasks):
    """Anuleaza un interviu si trimite email candidatului."""
    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        db.close()
        return JSONResponse({"error": "Interviul nu a fost gasit."}, status_code=404)

    cand_email = interview.candidate_email
    cand_name = interview.candidate_name or "Candidat"
    job_title = interview.job_title or "postul aplicat"
    data = interview.data
    ora = interview.ora
    locatie = interview.locatie or ""
    cand_id = interview.candidate_id

    interview.status = "anulat"
    db.commit()
    db.close()

    update_candidate_status(
        cand_id,
        "APLICAT",
        event_title="Interviu anulat",
        event_description=f"Interviul din {data} la {ora} a fost anulat de HR.",
    )

    if cand_email:
        background_tasks.add_task(
            _send_interview_cancelled_email,
            to_email=cand_email,
            candidate_name=cand_name,
            job_title=job_title,
            data=data,
            ora=ora,
        )

    background_tasks.add_task(
        _notify_hr_interview_event,
        "anulat_hr",
        cand_name,
        job_title,
        data,
        ora,
        locatie,
        cand_email or "",
    )

    return JSONResponse({"success": True})


@router.post("/calendar/api/interview/{interview_id}/edit-ora")
async def edit_interview_ora(
    background_tasks: BackgroundTasks,
    interview_id: int,
    ora: str = Form(...),
    data: str = Form(""),
):
    """Editeaza ora (si optional data) unui interviu."""
    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        db.close()
        return JSONResponse({"error": "Interviul nu a fost gasit."}, status_code=404)

    interview.ora = ora
    if data:
        interview.data = data
    db.commit()

    cand_name = interview.candidate_name or "Candidat"
    cand_email = interview.candidate_email or ""
    job_title = interview.job_title or "postul aplicat"
    new_data = interview.data
    new_ora = interview.ora
    locatie = interview.locatie or ""

    result = {
        "id": interview.id,
        "candidate_id": interview.candidate_id,
        "candidate_name": cand_name,
        "job_title": job_title,
        "data": new_data,
        "ora": new_ora,
        "locatie": locatie,
        "durata": interview.durata or 60,
        "status": interview.status,
    }
    db.close()

    # Notifica HR despre reprogramare
    background_tasks.add_task(
        _notify_hr_interview_event,
        "reprogramat_hr",
        cand_name,
        job_title,
        new_data,
        new_ora,
        locatie,
        cand_email,
    )

    # Notifica candidatul despre noua ora
    if cand_email:
        background_tasks.add_task(
            _send_interview_rescheduled_email,
            to_email=cand_email,
            candidate_name=cand_name,
            job_title=job_title,
            data=new_data,
            ora=new_ora,
            locatie=locatie,
        )

    return JSONResponse({"success": True, "interview": result})


def _send_interview_scheduled_email(
    to_email: str,
    candidate_name: str,
    job_title: str,
    data: str,
    ora: str,
    locatie: str,
):
    """Trimite email de confirmare interviu cu data, ora si locatia."""
    try:
        # Formatam data frumos: "2026-06-10" → "10 Iunie 2026"
        from datetime import datetime
        luni = {
            "01": "Ianuarie", "02": "Februarie", "03": "Martie",
            "04": "Aprilie", "05": "Mai", "06": "Iunie",
            "07": "Iulie", "08": "August", "09": "Septembrie",
            "10": "Octombrie", "11": "Noiembrie", "12": "Decembrie"
        }
        try:
            parts = data.split("-")
            data_formatata = f"{int(parts[2])} {luni.get(parts[1], parts[1])} {parts[0]}"
        except Exception:
            data_formatata = data

        locatie_line = f"\nLocatia / Link: {locatie}" if locatie else ""

        send_email_api(
            to_email=to_email,
            subject=f"Confirmare interviu — {job_title}",
            body=f"""Buna ziua, {candidate_name},

Va informam ca ati fost selectat(a) pentru un interviu pentru postul de {job_title}.

Data: {data_formatata}
Ora: {ora}{locatie_line}

Va rugam sa confirmati participarea raspunzand la acest email.

In caz ca nu puteti participa, va rugam sa ne anuntati cat mai curand posibil.

Cu respect,
Echipa HR NEXAS
"""
        )
    except Exception as e:
        print(f"[CALENDAR] WARN email interviu esuat catre {to_email}: {e}")


# ─── SELF-SCHEDULING: candidatul isi alege singur ora ───────────────────────


@router.post("/candidat/{candidate_id}/invita-self-schedule")
async def invite_self_schedule(candidate_id: int, background_tasks: BackgroundTasks, request: Request):
    """Genereaza un token unic si trimite email candidatului cu link de programare."""
    candidate = get_candidate_by_id(candidate_id)
    if not candidate:
        return JSONResponse({"error": "Candidatul nu a fost gasit."}, status_code=404)
    if not candidate.email:
        return JSONResponse({"error": "Candidatul nu are email."}, status_code=400)

    token_str = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(days=7)

    db = SessionLocal()
    tok = SchedulingToken(
        token=token_str,
        candidate_id=candidate_id,
        expires_at=expires,
    )
    db.add(tok)
    db.commit()
    db.close()

    base_url = str(request.base_url).rstrip("/")
    link = f"{base_url}/programeaza/{token_str}"

    background_tasks.add_task(
        _send_scheduling_invite_email,
        to_email=candidate.email,
        candidate_name=candidate.name or "Candidat",
        job_title=candidate.job_title or "postul pentru care ati aplicat",
        link=link,
    )

    return JSONResponse({"success": True, "message": "Invitatie trimisa cu succes."})


@router.get("/programeaza/{token}", response_class=HTMLResponse)
async def scheduling_page(request: Request, token: str):
    """Pagina publica unde candidatul isi alege data si ora interviului."""
    db = SessionLocal()
    tok = db.query(SchedulingToken).filter(SchedulingToken.token == token).first()
    db.close()

    if not tok:
        return HTMLResponse("<h2>Link invalid.</h2>", status_code=404)
    if tok.expires_at < datetime.utcnow():
        return HTMLResponse("<h2>Link-ul a expirat.</h2>", status_code=410)
    if tok.used:
        html = _public_env.get_template("schedule.html").render(
            token=token, candidate_name="", job_title="", already_used=True,
        )
        return HTMLResponse(html)

    db = SessionLocal()
    candidate = db.query(Candidate).filter(Candidate.id == tok.candidate_id).first()
    db.close()

    html = _public_env.get_template("schedule.html").render(
        token=token,
        candidate_name=candidate.name if candidate else "Candidat",
        job_title=candidate.job_title if candidate else "",
    )
    return HTMLResponse(html)


@router.get("/api/sloturi/{token}")
async def get_available_slots(token: str, data: str):
    """Returneaza sloturile orare disponibile (09:00-16:00) pt data data, minus cele ocupate."""
    db = SessionLocal()
    tok = db.query(SchedulingToken).filter(SchedulingToken.token == token).first()
    if not tok or tok.used or tok.expires_at < datetime.utcnow():
        db.close()
        return JSONResponse({"error": "Token invalid."}, status_code=403)

    booked = db.query(Interview.ora).filter(Interview.data == data).all()
    db.close()

    booked_hours = {row.ora for row in booked}
    all_slots = [f"{h:02d}:00" for h in range(9, 17)]
    available = [s for s in all_slots if s not in booked_hours]

    return JSONResponse({"slots": available})


@router.post("/programeaza/{token}")
async def submit_schedule(
    token: str,
    background_tasks: BackgroundTasks,
    request: Request,
    data: str = Form(...),
    ora: str = Form(...),
    locatie: str = Form("Sediu Firma NEXAS"),
):
    """Candidatul trimite formularul — salvam interviul si notificam HR-ul."""
    db = SessionLocal()
    tok = db.query(SchedulingToken).filter(SchedulingToken.token == token).first()

    if not tok:
        db.close()
        return HTMLResponse("<h2>Link invalid.</h2>", status_code=404)
    if tok.used:
        db.close()
        html = _public_env.get_template("schedule.html").render(
            token=token, candidate_name="", job_title="", already_used=True,
        )
        return HTMLResponse(html)
    if tok.expires_at < datetime.utcnow():
        db.close()
        return HTMLResponse("<h2>Link-ul a expirat.</h2>", status_code=410)

    candidate = db.query(Candidate).filter(Candidate.id == tok.candidate_id).first()
    if not candidate:
        db.close()
        return HTMLResponse("<h2>Candidatul nu a fost gasit.</h2>", status_code=404)

    # Extragem valorile inainte de a inchide sesiunea
    cand_id = candidate.id
    cand_name = candidate.name or "Candidat"
    cand_email = candidate.email
    cand_job = candidate.job_title or "—"

    interview = Interview(
        candidate_id=cand_id,
        candidate_name=candidate.name,
        candidate_email=cand_email,
        job_title=candidate.job_title,
        data=data,
        ora=ora,
        locatie=locatie,
        durata=60,
        status="programat",
        action_token=secrets.token_urlsafe(24),
    )
    db.add(interview)

    tok.used = True
    db.commit()
    db.close()

    update_candidate_status(
        cand_id,
        "INTERVIU",
        event_title="Interviu programat (self-schedule)",
        event_description=f"Candidatul a ales: {data} la {ora}. Locatie: {locatie}.",
    )

    background_tasks.add_task(
        _notify_hr_interview_event,
        "programat_candidat",
        cand_name,
        cand_job,
        data,
        ora,
        locatie,
        cand_email or "",
    )

    html = _public_env.get_template("schedule.html").render(
        token=token,
        candidate_name=cand_name,
        job_title=cand_job if cand_job != "—" else "",
        confirmed=True,
        data=data,
        ora=ora,
        locatie=locatie,
    )
    return HTMLResponse(html)


def _send_scheduling_invite_email(to_email: str, candidate_name: str, job_title: str, link: str):
    """Trimite email candidatului cu link-ul de self-scheduling."""
    try:
        send_email_api(
            to_email=to_email,
            subject=f"Programeaza-ti interviul — {job_title}",
            body=f"""Buna ziua, {candidate_name},

Va felicitam! Ati fost selectat(a) pentru un interviu pentru postul de {job_title}.

Va rugam sa va programati interviul accesand link-ul de mai jos si alegand data si ora care va convine:

PROGRAMEAZA INTERVIU:
{link}

Link-ul este valabil 7 zile.

Cu respect,
Echipa HR NEXAS
""",
        )
    except Exception as e:
        print(f"[SCHEDULE] WARN email invite esuat catre {to_email}: {e}")


def _notify_hr_interview_event(
    event_type: str,
    candidate_name: str,
    job_title: str,
    data: str,
    ora: str,
    locatie: str = "",
    candidate_email: str = "",
):
    """Notifica HR-ul pentru orice eveniment legat de un interviu.

    event_type: 'programat_candidat' | 'confirmat_candidat' | 'anulat_candidat'
                | 'programat_hr' | 'anulat_hr' | 'reprogramat_hr'
    """
    try:
        hr_email = os.getenv("HR_NOTIFICATION_EMAIL") or "nexas.axs@gmail.com"

        luni = {
            "01": "Ianuarie", "02": "Februarie", "03": "Martie",
            "04": "Aprilie", "05": "Mai", "06": "Iunie",
            "07": "Iulie", "08": "August", "09": "Septembrie",
            "10": "Octombrie", "11": "Noiembrie", "12": "Decembrie"
        }
        try:
            parts = data.split("-")
            data_fmt = f"{int(parts[2])} {luni.get(parts[1], parts[1])} {parts[0]}"
        except Exception:
            data_fmt = data

        subjects = {
            "programat_candidat": f"S-A PROGRAMAT — {candidate_name.upper()} — {data_fmt} {ora}",
            "confirmat_candidat": f"CONFIRMAT — {candidate_name.upper()} — {data_fmt} {ora}",
            "anulat_candidat":    f"ANULAT DE CANDIDAT — {candidate_name.upper()} — {data_fmt} {ora}",
            "programat_hr":       f"INTERVIU PROGRAMAT — {candidate_name.upper()} — {data_fmt} {ora}",
            "anulat_hr":          f"INTERVIU ANULAT DE HR — {candidate_name.upper()} — {data_fmt} {ora}",
            "reprogramat_hr":     f"INTERVIU REPROGRAMAT — {candidate_name.upper()} — {data_fmt} {ora}",
        }
        messages = {
            "programat_candidat": f"Candidatul {candidate_name} si-a ales singur data si ora interviului (self-schedule).",
            "confirmat_candidat": f"Candidatul {candidate_name} a CONFIRMAT participarea la interviu.",
            "anulat_candidat":    f"Candidatul {candidate_name} a ANULAT interviul din link-ul de email.",
            "programat_hr":       f"Ai programat un interviu nou pentru {candidate_name}.",
            "anulat_hr":          f"Interviul candidatului {candidate_name} a fost ANULAT de HR.",
            "reprogramat_hr":     f"Interviul candidatului {candidate_name} a fost REPROGRAMAT de HR.",
        }
        subject = subjects.get(event_type, f"Eveniment interviu — {candidate_name.upper()}")
        mesaj = messages.get(event_type, f"Eveniment: {event_type}")

        locatie_line = f"\nLocatie: {locatie}" if locatie else ""
        email_line = f"\nEmail candidat: {candidate_email}" if candidate_email else ""

        send_email_api(
            to_email=hr_email,
            subject=subject,
            body=f"""{mesaj}

Candidat: {candidate_name}
Post: {job_title}
Data: {data_fmt}
Ora: {ora}{locatie_line}{email_line}

Poti vedea detaliile in platforma NEXAS HR.

Echipa NEXAS HR
""",
        )
    except Exception as e:
        print(f"[HR-NOTIF] WARN email HR notif '{event_type}' esuat: {e}")


def _send_hr_notification_email(candidate_name: str, job_title: str, data: str, ora: str, locatie: str):
    """Notifica HR-ul cand un candidat si-a programat interviul."""
    try:
        hr_email = os.getenv("HR_NOTIFICATION_EMAIL") or "nexas.axs@gmail.com"

        luni = {
            "01": "Ianuarie", "02": "Februarie", "03": "Martie",
            "04": "Aprilie", "05": "Mai", "06": "Iunie",
            "07": "Iulie", "08": "August", "09": "Septembrie",
            "10": "Octombrie", "11": "Noiembrie", "12": "Decembrie"
        }
        try:
            parts = data.split("-")
            data_fmt = f"{int(parts[2])} {luni.get(parts[1], parts[1])} {parts[0]}"
        except Exception:
            data_fmt = data

        send_email_api(
            to_email=hr_email,
            subject=f"{candidate_name.upper()} - PROGRAMAT LA INTERVIU {data_fmt} {ora}",
            body=f"""Un candidat si-a programat interviul.

Candidat: {candidate_name}
Post: {job_title}
Data: {data_fmt}
Ora: {ora}
Locatie: {locatie}

Poti vedea detaliile in platforma NEXAS HR.

Echipa NEXAS HR
""",
        )
    except Exception as e:
        print(f"[SCHEDULE] WARN email HR notif esuat: {e}")


def _send_interview_cancelled_email(
    to_email: str,
    candidate_name: str,
    job_title: str,
    data: str,
    ora: str,
):
    """Notifica candidatul ca interviul sau a fost anulat de HR."""
    try:
        luni = {
            "01": "Ianuarie", "02": "Februarie", "03": "Martie",
            "04": "Aprilie", "05": "Mai", "06": "Iunie",
            "07": "Iulie", "08": "August", "09": "Septembrie",
            "10": "Octombrie", "11": "Noiembrie", "12": "Decembrie"
        }
        try:
            parts = data.split("-")
            data_fmt = f"{int(parts[2])} {luni.get(parts[1], parts[1])} {parts[0]}"
        except Exception:
            data_fmt = data

        send_email_api(
            to_email=to_email,
            subject=f"Interviu anulat — {job_title}",
            body=f"""Buna ziua, {candidate_name},

Va informam ca interviul programat pentru postul de {job_title} in data de {data_fmt} la ora {ora} a fost anulat.

Va rugam sa ne contactati pentru a stabili o noua data, daca sunteti in continuare interesat(a) de aceasta pozitie.

Cu respect,
Echipa HR NEXAS
""",
        )
    except Exception as e:
        print(f"[SCHEDULE] WARN email anulare esuat catre {to_email}: {e}")


def _send_interview_rescheduled_email(
    to_email: str,
    candidate_name: str,
    job_title: str,
    data: str,
    ora: str,
    locatie: str = "",
):
    """Notifica candidatul ca ora/data interviului a fost modificata de HR."""
    try:
        luni = {
            "01": "Ianuarie", "02": "Februarie", "03": "Martie",
            "04": "Aprilie", "05": "Mai", "06": "Iunie",
            "07": "Iulie", "08": "August", "09": "Septembrie",
            "10": "Octombrie", "11": "Noiembrie", "12": "Decembrie"
        }
        try:
            parts = data.split("-")
            data_fmt = f"{int(parts[2])} {luni.get(parts[1], parts[1])} {parts[0]}"
        except Exception:
            data_fmt = data

        locatie_line = f"\nLocatie: {locatie}" if locatie else ""

        send_email_api(
            to_email=to_email,
            subject=f"Interviu reprogramat — {job_title}",
            body=f"""Buna ziua, {candidate_name},

Va informam ca interviul pentru postul de {job_title} a fost reprogramat.

Noua data: {data_fmt}
Noua ora: {ora}{locatie_line}

Va rugam sa confirmati disponibilitatea. Ne cerem scuze pentru inconvenient.

Cu respect,
Echipa HR NEXAS
""",
        )
    except Exception as e:
        print(f"[SCHEDULE] WARN email reprogramare esuat catre {to_email}: {e}")


# ─── CONFIRMARE / ANULARE INTERVIU (linkuri din email) ───────────────────────

def _interview_action_page(title: str, message: str, color: str = "#22c7b8") -> str:
    return f"""<!DOCTYPE html>
<html lang="ro"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} | NEXAS HR</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;
background:linear-gradient(135deg,#121827,#171f33 55%,#1d263d);padding:24px}}
.card{{background:#1a2238;border:1px solid rgba(255,255,255,.12);border-radius:22px;padding:40px 36px;
max-width:440px;width:100%;text-align:center;box-shadow:0 8px 40px rgba(0,0,0,.3)}}
.icon{{font-size:52px;margin-bottom:18px}}
h2{{font-size:22px;font-weight:800;color:#fff;margin-bottom:10px}}
p{{font-size:14px;color:rgba(255,255,255,.65);line-height:1.6}}
.badge{{display:inline-block;margin-top:18px;padding:8px 20px;border-radius:99px;
font-size:12px;font-weight:700;background:rgba(255,255,255,.07);color:{color}}}
</style></head>
<body><div class="card">
<div class="icon">{'✅' if color == '#22c7b8' else '❌'}</div>
<h2>{title}</h2>
<p>{message}</p>
<div class="badge">NEXAS HR</div>
</div></body></html>"""


@router.get("/interviu/{action_token}/confirma", response_class=HTMLResponse)
async def confirma_interviu(action_token: str):
    """Candidatul confirma participarea la interviu din link-ul email."""
    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.action_token == action_token).first()
    if not interview:
        db.close()
        return HTMLResponse(_interview_action_page("Link invalid", "Acest link nu este valid.", "#ff6b70"), status_code=404)

    if interview.status == "anulat":
        db.close()
        return HTMLResponse(_interview_action_page(
            "Interviu anulat",
            "Interviul a fost deja anulat. Contacteaza-ne pentru reprogramare.",
            "#ff6b70"
        ))

    interview.status = "confirmat"
    db.commit()
    ora = interview.ora
    data = interview.data
    job = interview.job_title or "postul aplicat"
    cand_name = interview.candidate_name or "Candidat"
    cand_email = interview.candidate_email or ""
    locatie = interview.locatie or ""
    db.close()

    try:
        _notify_hr_interview_event(
            "confirmat_candidat", cand_name, job, data, ora,
            locatie=locatie, candidate_email=cand_email,
        )
    except Exception:
        pass

    return HTMLResponse(_interview_action_page(
        "Participare confirmata!",
        f"Multumim! Te asteptam pe {data} la ora {ora} pentru interviul de {job}.",
        "#22c7b8"
    ))


@router.get("/interviu/{action_token}/anuleaza", response_class=HTMLResponse)
async def anuleaza_interviu(action_token: str):
    """Candidatul anuleaza interviul din link-ul email."""
    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.action_token == action_token).first()
    if not interview:
        db.close()
        return HTMLResponse(_interview_action_page("Link invalid", "Acest link nu este valid.", "#ff6b70"), status_code=404)

    if interview.status in ("anulat", "finalizat"):
        db.close()
        return HTMLResponse(_interview_action_page(
            "Interviu deja procesat",
            "Interviul a fost deja anulat sau finalizat.",
            "#ff6b70"
        ))

    cand_id = interview.candidate_id
    cand_name = interview.candidate_name or "Candidat"
    cand_email = interview.candidate_email or ""
    job = interview.job_title or "postul aplicat"
    data = interview.data
    ora = interview.ora
    locatie = interview.locatie or ""
    interview.status = "anulat"
    db.commit()
    db.close()

    try:
        update_candidate_status(
            cand_id,
            "APLICAT",
            event_title="Interviu anulat de candidat",
            event_description="Candidatul a anulat interviul din link-ul de reminder.",
        )
    except Exception:
        pass

    try:
        _notify_hr_interview_event(
            "anulat_candidat", cand_name, job, data, ora,
            locatie=locatie, candidate_email=cand_email,
        )
    except Exception:
        pass

    return HTMLResponse(_interview_action_page(
        "Interviu anulat",
        "Interviul a fost anulat. Echipa HR va fi notificata si te vom contacta pentru reprogramare.",
        "#f59e0b"
    ))


# ── Interview Scorecard ───────────────────────────────────────────────────────

@router.get("/candidat/{candidate_id}/scorecard/{interview_id}", response_class=HTMLResponse)
async def scorecard_form(request: Request, candidate_id: int, interview_id: int):
    candidate = get_candidate_by_id(candidate_id)
    if not candidate:
        return RedirectResponse(url="/candidati", status_code=303)

    db = SessionLocal()
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    existing = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == candidate_id,
        InterviewScorecard.interview_id == interview_id,
    ).first()
    db.close()

    if not interview or interview.candidate_id != candidate_id:
        return RedirectResponse(url=f"/candidat/{candidate_id}", status_code=303)

    avg_score = calculate_average_score(existing) if existing else None
    message = request.query_params.get("message")

    return templates.TemplateResponse(
        request=request,
        name="scorecard_form.html",
        context={
            "request": request,
            "candidate": candidate,
            "interview": interview,
            "scorecard": existing,
            "avg_score": avg_score,
            "recommendations": VALID_RECOMMENDATIONS,
            "message": message,
        }
    )


@router.post("/candidat/{candidate_id}/scorecard/{interview_id}")
async def scorecard_submit(
    request: Request,
    candidate_id: int,
    interview_id: int,
    evaluator_name: str = Form(""),
    technical_skills_score: int = Form(...),
    attention_to_detail_score: int = Form(...),
    communication_score: int = Form(...),
    team_fit_score: int = Form(...),
    overall_recommendation: str = Form(...),
    comments: str = Form(""),
):
    errors = []
    for label, val in [
        ("Technical skills", technical_skills_score),
        ("Attention to detail", attention_to_detail_score),
        ("Communication", communication_score),
        ("Team fit", team_fit_score),
    ]:
        if not (1 <= val <= 5):
            errors.append(f"{label} score must be between 1 and 5.")

    if overall_recommendation not in VALID_RECOMMENDATIONS:
        errors.append("Invalid recommendation value.")

    if errors:
        candidate = get_candidate_by_id(candidate_id)
        db = SessionLocal()
        interview = db.query(Interview).filter(Interview.id == interview_id).first()
        existing = db.query(InterviewScorecard).filter(
            InterviewScorecard.candidate_id == candidate_id,
            InterviewScorecard.interview_id == interview_id,
        ).first()
        db.close()
        return templates.TemplateResponse(
            request=request,
            name="scorecard_form.html",
            context={
                "request": request,
                "candidate": candidate,
                "interview": interview,
                "scorecard": existing,
                "avg_score": None,
                "recommendations": VALID_RECOMMENDATIONS,
                "error": " ".join(errors),
            },
            status_code=422,
        )

    db = SessionLocal()
    existing = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == candidate_id,
        InterviewScorecard.interview_id == interview_id,
    ).first()

    if existing:
        existing.evaluator_name = evaluator_name.strip() or existing.evaluator_name
        existing.technical_skills_score = technical_skills_score
        existing.attention_to_detail_score = attention_to_detail_score
        existing.communication_score = communication_score
        existing.team_fit_score = team_fit_score
        existing.overall_recommendation = overall_recommendation
        existing.comments = comments.strip()
    else:
        sc = InterviewScorecard(
            candidate_id=candidate_id,
            interview_id=interview_id,
            evaluator_name=evaluator_name.strip() or None,
            technical_skills_score=technical_skills_score,
            attention_to_detail_score=attention_to_detail_score,
            communication_score=communication_score,
            team_fit_score=team_fit_score,
            overall_recommendation=overall_recommendation,
            comments=comments.strip() or None,
        )
        db.add(sc)

    db.commit()
    db.close()

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="SCORECARD",
        title="Scorecard completat",
        description=f"Evaluare interviu: {overall_recommendation}" + (f" – de catre {evaluator_name.strip()}" if evaluator_name.strip() else ""),
    )

    return RedirectResponse(
        url=f"/candidat/{candidate_id}/scorecard/{interview_id}?message=Scorecard salvat cu succes.",
        status_code=303,
    )


# ── Scorecard Token (link public pentru evaluator extern) ────────────────────

@router.post("/candidat/{candidate_id}/genereaza-link-evaluare")
async def genereaza_link_evaluare(
    request: Request,
    candidate_id: int,
    interview_id: int = Form(0),
    evaluator_name: str = Form(""),
    evaluator_email: str = Form(""),
):
    candidate = get_candidate_by_id(candidate_id)
    if not candidate:
        return JSONResponse({"success": False, "error": "Candidatul nu există."}, status_code=404)

    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(days=7)

    db = SessionLocal()
    sc_token = ScorecardToken(
        token=token,
        candidate_id=candidate_id,
        interview_id=interview_id if interview_id else None,
        evaluator_name=evaluator_name.strip() or None,
        evaluator_email=evaluator_email.strip() or None,
        expires_at=expires,
    )
    db.add(sc_token)
    db.commit()
    db.close()

    base_url = str(request.base_url).rstrip("/")
    link = f"{base_url}/evalueaza/{token}"

    log_candidate_event(
        candidate_id=candidate_id,
        event_type="SCORECARD_LINK",
        title="Link evaluare generat",
        description=f"Link trimis către: {evaluator_name.strip() or evaluator_email.strip() or 'evaluator extern'}",
    )

    return JSONResponse({"success": True, "link": link, "expires": expires.strftime("%d.%m.%Y")})


@router.get("/evalueaza/{token}", response_class=HTMLResponse)
async def evalueaza_get(request: Request, token: str):
    db = SessionLocal()
    sc_token = db.query(ScorecardToken).filter(ScorecardToken.token == token).first()
    db.close()

    if not sc_token:
        return HTMLResponse(_scorecard_token_error("Link invalid sau expirat."))

    if sc_token.expires_at < datetime.utcnow():
        return HTMLResponse(_scorecard_token_error("Acest link a expirat. Solicită un link nou de la HR."))

    candidate = get_candidate_by_id(sc_token.candidate_id)
    if not candidate:
        return HTMLResponse(_scorecard_token_error("Candidatul nu mai există în sistem."))

    db = SessionLocal()
    interview = None
    if sc_token.interview_id:
        interview = db.query(Interview).filter(Interview.id == sc_token.interview_id).first()
    existing = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == sc_token.candidate_id,
        InterviewScorecard.interview_id == sc_token.interview_id,
    ).first() if sc_token.interview_id else None
    db.close()

    already_submitted = sc_token.used
    avg_score = calculate_average_score(existing) if existing else None

    return templates.TemplateResponse(
        request=request,
        name="evalueaza_public.html",
        context={
            "request": request,
            "token": token,
            "candidate": candidate,
            "interview": interview,
            "scorecard": existing,
            "avg_score": avg_score,
            "already_submitted": already_submitted,
            "evaluator_name": sc_token.evaluator_name or "",
            "recommendations": VALID_RECOMMENDATIONS,
        }
    )


@router.post("/evalueaza/{token}")
async def evalueaza_post(
    request: Request,
    token: str,
    evaluator_name: str = Form(""),
    technical_skills_score: int = Form(...),
    attention_to_detail_score: int = Form(...),
    communication_score: int = Form(...),
    team_fit_score: int = Form(...),
    overall_recommendation: str = Form(...),
    comments: str = Form(""),
):
    db = SessionLocal()
    sc_token = db.query(ScorecardToken).filter(ScorecardToken.token == token).first()
    db.close()

    if not sc_token or sc_token.expires_at < datetime.utcnow():
        return HTMLResponse(_scorecard_token_error("Link invalid sau expirat."))

    errors = []
    for label, val in [
        ("Technical skills", technical_skills_score),
        ("Attention to detail", attention_to_detail_score),
        ("Communication", communication_score),
        ("Team fit", team_fit_score),
    ]:
        if not (1 <= val <= 5):
            errors.append(f"{label} trebuie să fie între 1 și 5.")

    if overall_recommendation not in VALID_RECOMMENDATIONS:
        errors.append("Recomandare invalidă.")

    if errors:
        candidate = get_candidate_by_id(sc_token.candidate_id)
        db = SessionLocal()
        interview = db.query(Interview).filter(Interview.id == sc_token.interview_id).first() if sc_token.interview_id else None
        db.close()
        return templates.TemplateResponse(
            request=request,
            name="evalueaza_public.html",
            context={
                "request": request,
                "token": token,
                "candidate": candidate,
                "interview": interview,
                "scorecard": None,
                "avg_score": None,
                "already_submitted": False,
                "evaluator_name": evaluator_name,
                "recommendations": VALID_RECOMMENDATIONS,
                "error": " ".join(errors),
            },
            status_code=422,
        )

    db = SessionLocal()
    existing = db.query(InterviewScorecard).filter(
        InterviewScorecard.candidate_id == sc_token.candidate_id,
        InterviewScorecard.interview_id == sc_token.interview_id,
    ).first() if sc_token.interview_id else None

    if existing:
        existing.evaluator_name = evaluator_name.strip() or existing.evaluator_name
        existing.technical_skills_score = technical_skills_score
        existing.attention_to_detail_score = attention_to_detail_score
        existing.communication_score = communication_score
        existing.team_fit_score = team_fit_score
        existing.overall_recommendation = overall_recommendation
        existing.comments = comments.strip()
    else:
        sc = InterviewScorecard(
            candidate_id=sc_token.candidate_id,
            interview_id=sc_token.interview_id,
            evaluator_name=evaluator_name.strip() or sc_token.evaluator_name,
            technical_skills_score=technical_skills_score,
            attention_to_detail_score=attention_to_detail_score,
            communication_score=communication_score,
            team_fit_score=team_fit_score,
            overall_recommendation=overall_recommendation,
            comments=comments.strip() or None,
        )
        db.add(sc)

    sc_token.used = True
    db.commit()
    db.close()

    evaluator_display = evaluator_name.strip() or sc_token.evaluator_name or "evaluator extern"
    log_candidate_event(
        candidate_id=sc_token.candidate_id,
        event_type="SCORECARD",
        title="Evaluare completată de evaluator extern",
        description=f"Recomandare: {overall_recommendation} — de către {evaluator_display}",
    )

    return RedirectResponse(url=f"/evalueaza/{token}?ok=1", status_code=303)


def _scorecard_token_error(msg: str) -> str:
    return f"""<!DOCTYPE html><html lang="ro" data-theme="dark">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Eroare | NEXAS HR</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;700;800&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;
background:linear-gradient(135deg,#121827,#171f33 50%,#1d263d);color:#fff;padding:20px}}
.box{{background:#1a2238;border:1px solid rgba(255,255,255,.09);border-radius:22px;padding:40px;max-width:440px;text-align:center}}
.icon{{font-size:48px;margin-bottom:16px}}
h2{{font-size:20px;font-weight:800;margin-bottom:10px}}
p{{font-size:14px;color:rgba(255,255,255,.6);line-height:1.6}}
</style></head>
<body><div class="box"><div class="icon">⚠️</div><h2>Link invalid</h2><p>{msg}</p></div></body></html>"""
