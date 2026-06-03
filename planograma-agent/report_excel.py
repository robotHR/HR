import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLORS = {
    "A": ("FF6B6B", "FFFFFF"),
    "B": ("4DABF7", "000000"),
    "C": ("69DB7C", "000000"),
    "header": ("1e1e2e", "00d4ff"),
}


def _header_style(ws, row: int, cols: int, bg: str, fg: str):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def generate(result: dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Sumar ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sumar"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20

    ws1.append(["Raport Optimizare Planogramă"])
    ws1["A1"].font = Font(bold=True, size=14, color="00d4ff")
    ws1.append([f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
    ws1.append([])

    stats = result["stats"]
    rows = [
        ("Total SKU-uri", stats["total_skus"]),
        ("SKU Clasa A", f"{stats['a_count']} ({stats['a_pct']}%)"),
        ("SKU Clasa B", f"{stats['b_count']} ({stats['b_pct']}%)"),
        ("SKU Clasa C", f"{stats['c_count']} ({stats['c_pct']}%)"),
        ("Volum Clasa A", f"{stats['a_vol_pct']}%"),
        ("Volum Clasa B", f"{stats['b_vol_pct']}%"),
        ("Volum Clasa C", f"{stats['c_vol_pct']}%"),
        ("Comenzi/zi", stats["orders_per_day"]),
        ("Angajați depozit", stats["employees"]),
    ]
    for r in rows:
        ws1.append(list(r))

    ws1.append([])
    ws1.append(["Analiză AI (Narativ)"])
    ws1[f"A{ws1.max_row}"].font = Font(bold=True, color="00d4ff")
    narrative = result.get("narrative", "Indisponibil")
    ws1.append([narrative])
    ws1[f"A{ws1.max_row}"].alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[ws1.max_row].height = 200

    # ── Sheet 2: Recomandari_SKU ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Recomandari_SKU")
    headers = ["Cod", "Produs", "Clasă", "Zonă Curentă", "Zonă Recomandată", "Acțiune", "Volum", "Score", "Status"]
    ws2.append(headers)
    _header_style(ws2, 1, len(headers), "1e1e2e", "00d4ff")

    col_widths = [12, 40, 8, 20, 22, 15, 12, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for rec in result["recommendations"]:
        row = [
            rec["cod"], rec["produs"], rec["clasa"],
            rec["zona_curenta"], rec["zona_recomandata"],
            rec["actiune"], rec["volum"], rec["score"], rec["status"],
        ]
        ws2.append(row)
        r = ws2.max_row
        cls = rec["clasa"]
        if cls in COLORS:
            bg, fg = COLORS[cls]
            fill = PatternFill("solid", fgColor=bg)
            ws2.cell(r, 3).fill = fill
            ws2.cell(r, 3).font = Font(color=fg, bold=True)

    # ── Sheet 3: Top_20 ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Top_20")
    t20_headers = ["#", "Cod", "Produs", "Clasă", "Volum Total", "Score"]
    ws3.append(t20_headers)
    _header_style(ws3, 1, len(t20_headers), "1e1e2e", "00d4ff")
    ws3.column_dimensions["C"].width = 40

    recs_sorted = sorted(result["recommendations"], key=lambda x: x["volum"], reverse=True)[:20]
    for i, rec in enumerate(recs_sorted, 1):
        ws3.append([i, rec["cod"], rec["produs"], rec["clasa"], rec["volum"], rec["score"]])
        r = ws3.max_row
        cls = rec["clasa"]
        if cls in COLORS:
            bg, fg = COLORS[cls]
            ws3.cell(r, 4).fill = PatternFill("solid", fgColor=bg)
            ws3.cell(r, 4).font = Font(color=fg, bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
